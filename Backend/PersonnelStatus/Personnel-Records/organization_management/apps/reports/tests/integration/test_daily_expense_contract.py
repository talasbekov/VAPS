import pytest
from datetime import date, timedelta
from django.utils import timezone
import io
from openpyxl import load_workbook

from organization_management.apps.divisions.models import Division
from organization_management.apps.staff_unit.models import StaffUnit
from organization_management.apps.employees.models import Employee
from organization_management.apps.statuses.models import EmployeeStatus
from organization_management.apps.reports.models import Report
from django.contrib.auth import get_user_model

User = get_user_model()

@pytest.fixture(autouse=True)
def mock_template_path(tmp_path, monkeypatch):
    """
    Создает временный валидный xlsx шаблон и подменяет settings.BASE_DIR,
    чтобы код utils.py искал файл в правильном месте.
    """
    import os
    from openpyxl import Workbook
    from django.conf import settings

    reports_dir = tmp_path / "apps" / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    template_file = reports_dir / "расход.xlsx"

    wb = Workbook()
    wb.save(template_file)
    monkeypatch.setattr(settings, 'BASE_DIR', str(tmp_path))

    return str(template_file)

@pytest.fixture
def structure_setup():
    department = Division.objects.create(name="Главный Департамент", code="root", division_type=Division.DivisionType.DEPARTMENT)
    directorate1 = Division.objects.create(name="Управление 1", code="dir1", division_type=Division.DivisionType.DIRECTORATE, parent=department)
    department.refresh_from_db()
    directorate1.refresh_from_db()

    division1 = Division.objects.create(name="Отдел 1", code="div1", division_type=Division.DivisionType.DIVISION, parent=directorate1)
    division1.refresh_from_db()

    su_dept_1 = StaffUnit.objects.create(division=department, index=1)
    su_dir1_1 = StaffUnit.objects.create(division=directorate1, index=1)
    su_dir1_2 = StaffUnit.objects.create(division=directorate1, index=2)
    su_dir1_3 = StaffUnit.objects.create(division=directorate1, index=3)
    su_dir1_vacant = StaffUnit.objects.create(division=directorate1, index=4)
    su_div1_1 = StaffUnit.objects.create(division=division1, index=1)

    emp_dept = Employee.objects.create(personnel_number="001", last_name="Иванов", first_name="Иван", gender=Employee.Gender.MALE); su_dept_1.employee = emp_dept; su_dept_1.save()
    emp_dir1_1 = Employee.objects.create(personnel_number="002", last_name="Петров", first_name="Петр", gender=Employee.Gender.MALE); su_dir1_1.employee = emp_dir1_1; su_dir1_1.save()
    emp_dir1_2 = Employee.objects.create(personnel_number="003", last_name="Сидоров", first_name="Сидор", gender=Employee.Gender.MALE); su_dir1_2.employee = emp_dir1_2; su_dir1_2.save()
    emp_dir1_3 = Employee.objects.create(personnel_number="004", last_name="Смирнов", first_name="Алексей", gender=Employee.Gender.MALE); su_dir1_3.employee = emp_dir1_3; su_dir1_3.save()
    emp_div1_1 = Employee.objects.create(personnel_number="005", last_name="Николаев", first_name="Николай", gender=Employee.Gender.MALE); su_div1_1.employee = emp_div1_1; su_div1_1.save()

    today = timezone.now().date()
    yesterday = today - timedelta(days=1)
    tomorrow = today + timedelta(days=1)

    test_user, _ = User.objects.get_or_create(username='status_creator', defaults={'password': 'password'})

    EmployeeStatus.objects.create(
        employee=emp_dept, status_type=EmployeeStatus.StatusType.IN_SERVICE,
        state=EmployeeStatus.StatusState.ACTIVE, start_date=yesterday, created_by=test_user
    )
    EmployeeStatus.objects.create(
        employee=emp_dir1_1, status_type=EmployeeStatus.StatusType.IN_SERVICE,
        state=EmployeeStatus.StatusState.ACTIVE, start_date=yesterday, created_by=test_user
    )
    EmployeeStatus.objects.create(
        employee=emp_dir1_2, status_type=EmployeeStatus.StatusType.VACATION,
        state=EmployeeStatus.StatusState.ACTIVE, start_date=yesterday, end_date=tomorrow, created_by=test_user
    )
    EmployeeStatus.objects.create(
        employee=emp_dir1_3, status_type=EmployeeStatus.StatusType.SICK_LEAVE,
        state=EmployeeStatus.StatusState.ACTIVE, start_date=yesterday, end_date=tomorrow, created_by=test_user
    )
    EmployeeStatus.objects.create(
        employee=emp_div1_1, status_type=EmployeeStatus.StatusType.BUSINESS_TRIP,
        state=EmployeeStatus.StatusState.ACTIVE, start_date=yesterday, end_date=tomorrow, created_by=test_user
    )

    return {
        'department': department,
        'directorate1': directorate1,
    }


@pytest.mark.django_db
class TestDailyExpenseSyncAsyncContract:
    def test_sync_vs_async_gap(self, structure_setup):
        """
        Проверяет расхождения между генератором Excel (Sync) и DataAggregator (Async).
        """
        from organization_management.apps.reports.utils import generate_personnel_expense_report
        from organization_management.apps.reports.infrastructure.data_aggregator import DataAggregator

        department = structure_setup['department']
        dir1 = structure_setup['directorate1']

        # 1. СИНХРОННЫЙ ПУТЬ (XLSX)
        file_buffer, _ = generate_personnel_expense_report(department.id)
        wb = load_workbook(io.BytesIO(file_buffer.read()))
        ws = wb.active

        sync_dir1 = {}
        sync_total = {}

        for row in ws.iter_rows(values_only=True):
            if row[0] == "Управление 1":
                sync_dir1 = {
                    "staff_units": row[1],
                    "employees": row[2],
                    "in_service": row[3],
                    "vacancies": row[4],
                    "vacation": row[5],
                    "trip": row[6],
                    "sick": row[7],
                }
            elif row[0] == "ИТОГО":
                sync_total = {
                    "staff_units": row[1],
                    "employees": row[2],
                    "in_service": row[3],
                    "vacancies": row[4],
                    "vacation": row[5],
                    "trip": row[6],
                    "sick": row[7],
                }

        assert sync_dir1, "Управление 1 row was not found in sync XLSX"
        assert sync_total, "ИТОГО row was not found in sync XLSX"

        assert sync_dir1 == {
            "staff_units": 5,
            "employees": 4,
            "in_service": 1,
            "vacancies": 1,
            "vacation": 1,
            "trip": 1,
            "sick": 1,
        }
        assert sync_total == {
            "staff_units": 6,
            "employees": 5,
            "in_service": 2,
            "vacancies": 1,
            "vacation": 1,
            "trip": 1,
            "sick": 1,
        }

        # 2. АСИНХРОННЫЙ ПУТЬ (DataAggregator)
        test_user, _ = User.objects.get_or_create(username='report_creator')
        report = Report.objects.create(
            job_id="test-job",
            report_type=Report.ReportType.PERSONNEL_ROSTER,
            division=department,
            created_by=test_user,
            date_from=timezone.now().date(),
            date_to=timezone.now().date(),
        )

        # We need to catch DataAggregator crashing because Employee model has no `division_id`.
        # DataAggregator currently uses: Employee.objects.filter(division_id__in=division_ids)
        # However Employee division is via StaffUnit: `employee.staff_unit.division_id`.
        aggregator = DataAggregator()
        async_data = aggregator.collect_data(report)

        # Map Async values
        def get_async_stats(division_id):
            for row in async_data['rows']:
                if row.get('division_id') == division_id:
                    return row
            return None

        async_dir1_raw = get_async_stats(dir1.id)

        async_dir1 = {
            "staff_units": async_dir1_raw['staff_unit'] if async_dir1_raw else 0,
            "employees": async_dir1_raw['total_working'] if async_dir1_raw else 0,
            "vacancies": async_dir1_raw['vacancies'] if async_dir1_raw else 0,
            "in_service": async_dir1_raw['in_service'] if async_dir1_raw else 0,
            "vacation": async_dir1_raw['vacation'] if async_dir1_raw else 0,
            "trip": async_dir1_raw['business_trip'] if async_dir1_raw else 0,
            "sick": async_dir1_raw['sick_leave'] if async_dir1_raw else 0,
        }

        async_total_raw = async_data.get('summary', {})
        async_total = {
            "staff_units": async_total_raw.get('staffing_qty', 0),
            "employees": async_total_raw.get('total_working', 0),
            "vacancies": async_total_raw.get('vacancies', 0),
            "in_service": async_total_raw.get('in_service', 0),
            "vacation": async_total_raw.get('vacation', 0),
            "trip": async_total_raw.get('business_trip', 0),
            "sick": async_total_raw.get('sick_leave', 0),
        }

        # Now they should match perfectly!
        assert sync_dir1["staff_units"] == async_dir1["staff_units"]
        assert sync_dir1["employees"] == async_dir1["employees"]
        assert sync_dir1["vacancies"] == async_dir1["vacancies"]
        assert sync_dir1["in_service"] == async_dir1["in_service"]
        assert sync_dir1["vacation"] == async_dir1["vacation"]
        assert sync_dir1["trip"] == async_dir1["trip"]
        assert sync_dir1["sick"] == async_dir1["sick"]

        assert sync_total["staff_units"] == async_total["staff_units"]
        assert sync_total["employees"] == async_total["employees"]
        assert sync_total["vacancies"] == async_total["vacancies"]
        assert sync_total["in_service"] == async_total["in_service"]
        assert sync_total["vacation"] == async_total["vacation"]
        assert sync_total["trip"] == async_total["trip"]
        assert sync_total["sick"] == async_total["sick"]

        # 4. Verify explicit text notes from generated sync XLSX (generated from DataAggregator now!)
        found_sidorov_vacation = False
        found_smirnov_sick = False
        found_nikolaev_trip = False

        notes_row = False
        for row in ws.iter_rows(values_only=True):
            if row[0] == "Управление 1":
                notes_row = True
                continue

            if notes_row:
                # The row immediately following the Directorate row contains the notes
                # Index 5 = Vacation notes (zero-indexed) -> Column 6 (F)
                if row[5] and "Сидоров Сидор" in str(row[5]):
                    found_sidorov_vacation = True

                # Index 6 = Trip notes -> Column 7 (G)
                if row[6] and "Николаев Николай" in str(row[6]):
                    found_nikolaev_trip = True

                # Index 7 = Sick notes -> Column 8 (H)
                if row[7] and "Смирнов Алексей" in str(row[7]):
                    found_smirnov_sick = True

                notes_row = False

        assert found_sidorov_vacation, "Сидоров is missing from vacation notes"
        assert found_smirnov_sick, "Смирнов is missing from sick notes"
        assert found_nikolaev_trip, "Николаев is missing from trip notes"
