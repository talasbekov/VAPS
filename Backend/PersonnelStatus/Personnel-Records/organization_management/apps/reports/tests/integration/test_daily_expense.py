import pytest
from datetime import date, timedelta
from django.utils import timezone
from django.urls import reverse
from rest_framework.test import APIClient
from openpyxl import load_workbook
import io

from organization_management.apps.divisions.models import Division
from organization_management.apps.staff_unit.models import StaffUnit
from organization_management.apps.employees.models import Employee
from organization_management.apps.statuses.models import EmployeeStatus
from django.contrib.auth import get_user_model

User = get_user_model()

@pytest.fixture(autouse=True)
def template_is_the_delivered_one():
    """Шаблон берётся ИЗ ПОСТАВКИ, а не создаётся пробой (Plane №254).

    Здесь стояла автофикстура `mock_template_path`: она клала во временную
    папку пустой xlsx и переставляла на неё `settings.BASE_DIR`. Из-за этого
    весь файл был зелёным, пока на живом стенде та же ручка отвечала 500 —
    файла `расход.xlsx` не было в репозитории ВООБЩЕ. Проба, которая чинит за
    код недостающий ресурс, перестаёт отвечать на вопрос «работает ли это в
    поставке», и именно этот класс слепоты снят.

    Теперь фикстура ничего не подменяет, а ПРОВЕРЯЕТ: шаблон обязан лежать по
    тому пути, куда смотрит `utils.generate_personnel_expense_report`. Нет
    файла — падает весь файл проб с внятной причиной, а не молчит.
    """
    import os
    from django.conf import settings

    path = os.path.join(settings.BASE_DIR, "apps/reports/расход.xlsx")
    assert os.path.exists(path), (
        f"Шаблона «расход.xlsx» нет в поставке ({path}). Пробы этого файла "
        f"проверяют сборку отчёта по НАСТОЯЩЕМУ шаблону и подменять его "
        f"пустым workbook'ом не должны — см. Plane №254."
    )
    return path


@pytest.fixture
def api_client():
    return APIClient()

@pytest.fixture
def test_user():
    user, _ = User.objects.get_or_create(username='test_user', defaults={'password': 'password'})
    return user

@pytest.fixture
def super_user():
    user, _ = User.objects.get_or_create(username='super_user', defaults={'password': 'password', 'is_superuser': True})
    return user

@pytest.fixture
def structure_setup():
    """
    Создает тестовую структуру:
    Департамент -> Управление -> Отдел
    С разными штатными единицами, сотрудниками и статусами.
    """
    # 1. Подразделения (Divisions)
    department = Division.objects.create(name="Главный Департамент", code="root", division_type=Division.DivisionType.DEPARTMENT)
    directorate1 = Division.objects.create(name="Управление 1", code="dir1", division_type=Division.DivisionType.DIRECTORATE, parent=department)
    department.refresh_from_db()
    directorate1.refresh_from_db()

    # Подчиненный отдел
    division1 = Division.objects.create(name="Отдел 1", code="div1", division_type=Division.DivisionType.DIVISION, parent=directorate1)
    division1.refresh_from_db()

    # 2. Штатные единицы
    su_dept_1 = StaffUnit.objects.create(division=department, index=1)
    su_dir1_1 = StaffUnit.objects.create(division=directorate1, index=1)
    su_dir1_2 = StaffUnit.objects.create(division=directorate1, index=2)
    su_dir1_3 = StaffUnit.objects.create(division=directorate1, index=3)
    su_dir1_vacant = StaffUnit.objects.create(division=directorate1, index=4)
    su_div1_1 = StaffUnit.objects.create(division=division1, index=1)

    # 3. Сотрудники
    emp_dept = Employee.objects.create(personnel_number="001", last_name="Иванов", first_name="Иван", gender=Employee.Gender.MALE); su_dept_1.employee = emp_dept; su_dept_1.save()
    emp_dir1_1 = Employee.objects.create(personnel_number="002", last_name="Петров", first_name="Петр", gender=Employee.Gender.MALE); su_dir1_1.employee = emp_dir1_1; su_dir1_1.save()
    emp_dir1_2 = Employee.objects.create(personnel_number="003", last_name="Сидоров", first_name="Сидор", gender=Employee.Gender.MALE); su_dir1_2.employee = emp_dir1_2; su_dir1_2.save()
    emp_dir1_3 = Employee.objects.create(personnel_number="004", last_name="Смирнов", first_name="Алексей", gender=Employee.Gender.MALE); su_dir1_3.employee = emp_dir1_3; su_dir1_3.save()
    emp_div1_1 = Employee.objects.create(personnel_number="005", last_name="Николаев", first_name="Николай", gender=Employee.Gender.MALE); su_div1_1.employee = emp_div1_1; su_div1_1.save()

    # 4. Статусы
    today = timezone.localdate()
    yesterday = today - timedelta(days=1)
    tomorrow = today + timedelta(days=1)

    test_user, _ = User.objects.get_or_create(username='status_creator', defaults={'password': 'password'})

    # Иванов: В строю
    EmployeeStatus.objects.create(
        employee=emp_dept,
        status_type=EmployeeStatus.StatusType.IN_SERVICE,
        state=EmployeeStatus.StatusState.ACTIVE,
        start_date=yesterday, created_by=test_user
    )

    # Петров: В строю
    EmployeeStatus.objects.create(
        employee=emp_dir1_1,
        status_type=EmployeeStatus.StatusType.IN_SERVICE,
        state=EmployeeStatus.StatusState.ACTIVE,
        start_date=yesterday, created_by=test_user
    )

    # Сидоров: В отпуске
    EmployeeStatus.objects.create(
        employee=emp_dir1_2,
        status_type=EmployeeStatus.StatusType.VACATION,
        state=EmployeeStatus.StatusState.ACTIVE,
        start_date=yesterday, created_by=test_user,
        end_date=tomorrow
    )

    # Смирнов: На больничном
    EmployeeStatus.objects.create(
        employee=emp_dir1_3,
        status_type=EmployeeStatus.StatusType.SICK_LEAVE,
        state=EmployeeStatus.StatusState.ACTIVE,
        start_date=yesterday, created_by=test_user,
        end_date=tomorrow
    )

    # Николаев: В командировке
    EmployeeStatus.objects.create(
        employee=emp_div1_1,
        status_type=EmployeeStatus.StatusType.BUSINESS_TRIP,
        state=EmployeeStatus.StatusState.ACTIVE,
        start_date=yesterday, created_by=test_user,
        end_date=tomorrow
    )

    return {
        'department': department,
        'directorate1': directorate1,
    }


@pytest.mark.django_db
class TestDailyExpenseIntegration:

    def test_expense_report_api_success(self, api_client, super_user, structure_setup):
        """
        Тестируем успешную генерацию отчета и его структуру через API.
        """
        api_client.force_authenticate(user=super_user)
        department = structure_setup['department']

        # Эндпоинт: /api/reports/reports/expense/<department_id>/
        url = reverse('report-expense', kwargs={'department_id': department.id})

        response = api_client.get(url)
        assert response.status_code == 200
        assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        # Проверяем содержимое файла
        wb = load_workbook(io.BytesIO(b"".join(response.streaming_content)))
        ws = wb.active

        # СТРОКА 1 (index 0 - header, index 1 - data)
        found_directorate = False
        for row in ws.iter_rows(values_only=True):
            if row[0] == "Управление 1":
                found_directorate = True
                assert row[1] == 5  # staff units count in directorate1 + division1 (su_dir1_1, su_dir1_2, su_dir1_3, su_dir1_vacant, su_div1_1)
                assert row[2] == 4  # employees count (Петров, Сидоров, Смирнов, Николаев)
                assert row[3] == 1  # in_service_count (Петров)
                assert row[4] == 1  # vacancies_count (su_dir1_vacant)
                assert row[5] == 1  # vacation_count (Сидоров)
                assert row[6] == 1  # trip (Николаев)
                assert row[7] == 1  # sick (Смирнов)

        assert found_directorate, "Управление 1 не найдено в отчете"

    def test_expense_report_function_direct(self, structure_setup):
        """
        Тестируем функцию generate_personnel_expense_report напрямую.
        Проверяем агрегацию и логику подсчета вакансий и сотрудников.
        """
        from organization_management.apps.reports.utils import generate_personnel_expense_report
        department = structure_setup['department']

        file_buffer, filename = generate_personnel_expense_report(department.id)

        # Проверяем структуру имени
        assert filename.startswith("расход_Главный Департамент_")
        assert filename.endswith(".xlsx")

        # Проверяем содержимое
        wb = load_workbook(io.BytesIO(file_buffer.read()))
        ws = wb.active

        # Ищем ИТОГО
        found_total = False
        for row in ws.iter_rows(values_only=True):
            if row[0] == "ИТОГО":
                found_total = True
                assert row[1] == 6  # total staff units (direct dept + all directorate descendants)
                assert row[2] == 5  # total employees (direct dept + all directorate descendants)
                assert row[3] == 2  # total in service (direct dept + all directorate descendants)
                assert row[4] == 1  # total vacancies
                assert row[5] == 1  # total vacation
                assert row[6] == 1  # total trip
                assert row[7] == 1  # total sick

        assert found_total, "Строка ИТОГО не найдена"

    def test_secondment_lands_on_the_receiving_side(self, structure_setup):
        """«Прикомандирован» считается там, КУДА человек пришёл.

        Колонка бралась по строкам «Прикомандирован из» и родному
        подразделению сотрудника — то есть человек числился прикомандированным
        к самому себе и попадал сразу в две колонки, «прикомандирован» и
        «откомандирован».
        """
        from organization_management.apps.reports.utils import (
            generate_personnel_expense_report,
        )
        from organization_management.apps.statuses.application.services import (
            StatusApplicationService,
        )

        department = structure_setup['department']
        directorate1 = structure_setup['directorate1']
        # Иванов состоит в департаменте (строка «Руководство»), уходит в
        # Управление 1 — принимающая сторона.
        ivanov = Employee.objects.get(personnel_number="001")
        author = User.objects.get(username='status_creator')
        today = timezone.localdate()
        StatusApplicationService().create_status(
            employee_id=ivanov.id,
            status_type=EmployeeStatus.StatusType.SECONDED_TO,
            start_date=today,
            end_date=today + timedelta(days=3),
            related_division_id=directorate1.id,
            user=author,
        )

        file_buffer, _filename = generate_personnel_expense_report(department.id)
        wb = load_workbook(io.BytesIO(file_buffer.read()))
        ws = wb.active

        directorate_row = None
        for row in ws.iter_rows(values_only=True):
            if row[0] == "Управление 1":
                directorate_row = row
        assert directorate_row is not None, "Управление 1 не найдено в отчете"

        # Колонка 12 — «Прикомандирован», колонка 13 — «Откомандирован».
        assert directorate_row[11] == 1
        assert directorate_row[12] == 0


@pytest.mark.django_db
def test_the_missing_template_answers_with_a_reason_and_not_a_500(
    django_user_model, tmp_path, monkeypatch
):
    """Отчёт без шаблона отказывает ВНЯТНО и называет рабочий путь.

    Найдено сквозной проверкой сценария заказчика (Plane №243): ручка отдавала
    500 с текстом системной ошибки наружу — «[Errno 2] No such file or
    directory», — а шаблона «расход.xlsx» нет не по пути, а вообще нигде в
    дереве. Молчаливая 500 хуже отсутствия ручки: её считают сломанной
    временно и возвращаются, вместо того чтобы пойти в сводку дня раздела ОМ,
    где документ действительно собирается.

    Красная на мутации: убери проверку существования шаблона — вернётся 500.
    """
    from django.conf import settings

    from organization_management.apps.divisions.models import Division

    # BASE_DIR переставляется на заведомо пустой каталог: проверяется ПОВЕДЕНИЕ
    # кода без шаблона, а не сегодняшнее состояние поставки. Само отсутствие
    # шаблона в поставке было дефектом и закрыто (№254, шаблон возвращён в
    # `apps/reports/`), но отказ обязан остаться внятным: файл можно потерять
    # снова при переезде каталогов или сборке образа, и тогда наружу опять
    # полезет «[Errno 2] No such file or directory».
    empty = tmp_path / "без-шаблона"
    empty.mkdir()
    monkeypatch.setattr(settings, "BASE_DIR", str(empty))

    user = django_user_model.objects.create_superuser(
        username="reports-probe", password="x", email="probe@example.org"
    )
    # APIClient с принудительной аутентификацией — как у соседних проб файла:
    # обычный `client.force_login` DRF-гейт этой ручки не проходит.
    api = APIClient()
    api.force_authenticate(user)
    department = Division.objects.create(
        name="Департамент пробы",
        code="DEP-REPORT-PROBE",
        division_type=Division.DivisionType.DEPARTMENT,
    )

    response = api.get(f"/api/reports/reports/expense/{department.pk}/")

    assert response.status_code == 400, response.content[:200]
    detail = response.json()["detail"]
    assert "расход.xlsx" in detail
    # Отказ обязан назвать, КУДА идти: иначе он честен, но бесполезен.
    assert "daily-summaries/export" in detail
