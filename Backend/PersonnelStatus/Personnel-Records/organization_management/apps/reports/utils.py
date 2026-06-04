"""
Утилиты для генерации отчетов
"""
import os
from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from django.conf import settings
from django.utils import timezone

from organization_management.apps.divisions.models import Division
from organization_management.apps.reports.infrastructure.data_aggregator import DataAggregator


def safe_set_cell_value(ws, row, col, value):
    """
    Безопасно устанавливает значение ячейки, даже если это объединенная ячейка.
    Записывает в первую (верхнюю левую) ячейку объединенного диапазона.
    """
    cell = ws.cell(row=row, column=col)

    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                min_row = merged_range.min_row
                min_col = merged_range.min_col
                ws.cell(row=min_row, column=min_col).value = value
                return
    else:
        cell.value = value


def generate_personnel_expense_report(department_id):
    """
    Генерирует отчет "Расход" по департаменту в памяти, используя DataAggregator (STORY-000.8).
    """
    template_path = os.path.join(
        settings.BASE_DIR,
        'apps/reports/расход.xlsx'
    )
    wb = load_workbook(template_path)
    ws = wb.active

    try:
        department = Division.objects.get(pk=department_id, division_type=Division.DivisionType.DEPARTMENT)
    except Division.DoesNotExist:
        raise ValueError(f'Департамент с ID {department_id} не найден')

    heads = Division.objects.filter(division_type=Division.DivisionType.DEPARTMENT, is_active=True).order_by('name')

    directorates = Division.objects.filter(
        parent=department,
        division_type=Division.DivisionType.DIRECTORATE,
        is_active=True
    ).order_by('name')

    # 1. СБОР ДАННЫХ (DataAggregator)
    aggregator = DataAggregator()
    ref_date = timezone.now().date()
    data = aggregator.collect_for_division(department, ref_date)

    head_metrics = data.get("head_metrics", {})
    summary = data.get("summary", {})

    # Ищем строки по directorate.id
    def get_dir_row(dir_id):
        for r in data.get("rows", []):
            if r["division_id"] == dir_id:
                return r
        return None

    # Итоговые счетчики (Берем напрямую из summary)
    total_staff_units = summary.get("staff_unit", 0)
    total_employees = summary.get("total_working", 0)
    total_in_service = summary.get("in_service", 0)
    total_vacancies = summary.get("vacancies", 0)
    total_vacation = summary.get("vacation", 0)
    total_sick = summary.get("sick_leave", 0)
    total_business_trip = summary.get("business_trip", 0)

    # Примечания ИТОГО
    total_on_duty = summary.get("notes", {}).get("on_duty_list", []) if "notes" in summary else []
    total_after_duty = summary.get("notes", {}).get("after_duty_list", []) if "notes" in summary else []
    total_training = summary.get("training", 0)
    total_seconded_from = summary.get("seconded_in", 0)
    total_seconded_to = summary.get("seconded_out", 0)

    # === ОБРАБОТКА РУКОВОДСТВА ДЕПАРТАМЕНТА (строки 4-5) ===
    head_row = 4

    # Извлекаем метрики руководителя
    head_staff_units = head_metrics.get("staff_unit", 0)
    head_employees = head_metrics.get("total_working", 0)
    head_in_service = head_metrics.get("in_service", 0)
    head_vacancies = head_metrics.get("vacancies", 0)
    head_vacation_count = head_metrics.get("vacation", 0)
    head_sick_count = head_metrics.get("sick_leave", 0)
    head_trip_count = head_metrics.get("business_trip", 0)

    head_on_duty_count = len(head_metrics.get("notes", {}).get("on_duty_list", []))
    head_after_duty_count = len(head_metrics.get("notes", {}).get("after_duty_list", []))
    head_training_count = head_metrics.get("training", 0)
    head_seconded_from_count = head_metrics.get("seconded_in", 0)
    head_seconded_to_count = head_metrics.get("seconded_out", 0)

    head_notes = head_metrics.get("notes", {})
    head_vacation_list = head_notes.get("vacation_list", [])
    head_trip_list = head_notes.get("trip_list", [])
    head_sick_list = head_notes.get("sick_list", [])
    head_on_duty_list = head_notes.get("on_duty_list", [])
    head_after_duty_list = head_notes.get("after_duty_list", [])
    head_training_list = head_notes.get("training_list", [])
    head_seconded_from_list = head_notes.get("seconded_from_list", [])
    head_seconded_to_list = head_notes.get("seconded_to_list", [])

    # СТРОКА 4: Руководство - Название + Числа
    safe_set_cell_value(ws, head_row, 1, department.name)
    safe_set_cell_value(ws, head_row, 2, head_staff_units)
    safe_set_cell_value(ws, head_row, 3, head_employees)
    safe_set_cell_value(ws, head_row, 4, head_in_service)
    safe_set_cell_value(ws, head_row, 5, head_vacancies)
    safe_set_cell_value(ws, head_row, 6, head_vacation_count)
    safe_set_cell_value(ws, head_row, 7, head_trip_count)
    safe_set_cell_value(ws, head_row, 8, head_sick_count)
    safe_set_cell_value(ws, head_row, 9, head_on_duty_count)
    safe_set_cell_value(ws, head_row, 10, head_after_duty_count)
    safe_set_cell_value(ws, head_row, 11, head_training_count)
    safe_set_cell_value(ws, head_row, 12, head_seconded_from_count)
    safe_set_cell_value(ws, head_row, 13, head_seconded_to_count)

    # СТРОКА 5: Руководство - Примечания
    head_notes_row = head_row + 1
    safe_set_cell_value(ws, head_notes_row, 6, "\n".join(head_vacation_list))
    safe_set_cell_value(ws, head_notes_row, 7, "\n".join(head_trip_list))
    safe_set_cell_value(ws, head_notes_row, 8, "\n".join(head_sick_list))
    safe_set_cell_value(ws, head_notes_row, 9, "\n".join(head_on_duty_list))
    safe_set_cell_value(ws, head_notes_row, 10, "\n".join(head_after_duty_list))
    safe_set_cell_value(ws, head_notes_row, 11, "\n".join(head_training_list))
    safe_set_cell_value(ws, head_notes_row, 12, "\n".join(head_seconded_from_list))
    safe_set_cell_value(ws, head_notes_row, 13, "\n".join(head_seconded_to_list))

    # === ОБРАБОТКА УПРАВЛЕНИЙ ===
    start_row = 6
    current_row = start_row

    for directorate in directorates:
        dr = get_dir_row(directorate.id)
        if not dr:
            continue

        staff_units_count = dr.get("staff_unit", 0)
        employees_count = dr.get("total_working", 0)
        in_service_count = dr.get("in_service", 0)
        vacancies_count = dr.get("vacancies", 0)
        vacation_count = dr.get("vacation", 0)
        trip_count = dr.get("business_trip", 0)
        sick_count = dr.get("sick_leave", 0)

        dr_notes = dr.get("notes", {})
        on_duty_list = dr_notes.get("on_duty_list", [])
        on_duty_count = len(on_duty_list)

        after_duty_list = dr_notes.get("after_duty_list", [])
        after_duty_count = len(after_duty_list)

        training_count = dr.get("training", 0)
        seconded_from_count = dr.get("seconded_in", 0)
        seconded_to_count = dr.get("seconded_out", 0)

        vacation_list = dr_notes.get("vacation_list", [])
        trip_list = dr_notes.get("trip_list", [])
        sick_list = dr_notes.get("sick_list", [])
        training_list = dr_notes.get("training_list", [])
        seconded_from_list = dr_notes.get("seconded_from_list", [])
        seconded_to_list = dr_notes.get("seconded_to_list", [])

        # Заполнение числовых данных (1 строка на управление)
        safe_set_cell_value(ws, current_row, 1, directorate.name)
        safe_set_cell_value(ws, current_row, 2, staff_units_count)
        safe_set_cell_value(ws, current_row, 3, employees_count)
        safe_set_cell_value(ws, current_row, 4, in_service_count)
        safe_set_cell_value(ws, current_row, 5, vacancies_count)
        safe_set_cell_value(ws, current_row, 6, vacation_count)
        safe_set_cell_value(ws, current_row, 7, trip_count)
        safe_set_cell_value(ws, current_row, 8, sick_count)
        safe_set_cell_value(ws, current_row, 9, on_duty_count)
        safe_set_cell_value(ws, current_row, 10, after_duty_count)
        safe_set_cell_value(ws, current_row, 11, training_count)
        safe_set_cell_value(ws, current_row, 12, seconded_from_count)
        safe_set_cell_value(ws, current_row, 13, seconded_to_count)

        # Заполнение примечаний ФИО
        notes_row = current_row + 1
        safe_set_cell_value(ws, notes_row, 6, "\n".join(vacation_list))
        safe_set_cell_value(ws, notes_row, 7, "\n".join(trip_list))
        safe_set_cell_value(ws, notes_row, 8, "\n".join(sick_list))
        safe_set_cell_value(ws, notes_row, 9, "\n".join(on_duty_list))
        safe_set_cell_value(ws, notes_row, 10, "\n".join(after_duty_list))
        safe_set_cell_value(ws, notes_row, 11, "\n".join(training_list))
        safe_set_cell_value(ws, notes_row, 12, "\n".join(seconded_from_list))
        safe_set_cell_value(ws, notes_row, 13, "\n".join(seconded_to_list))

        current_row += 2

    # === СТРОКА ИТОГО ===
    total_row = current_row

    # Totals for duty counts are now tracked correctly inside DataAggregator summary payload
    total_on_duty_count = summary.get("on_duty", 0)
    total_after_duty_count = summary.get("after_duty", 0)

    safe_set_cell_value(ws, total_row, 1, "ИТОГО")
    safe_set_cell_value(ws, total_row, 2, total_staff_units)
    safe_set_cell_value(ws, total_row, 3, total_employees)
    safe_set_cell_value(ws, total_row, 4, total_in_service)
    safe_set_cell_value(ws, total_row, 5, total_vacancies)
    safe_set_cell_value(ws, total_row, 6, total_vacation)
    safe_set_cell_value(ws, total_row, 7, total_business_trip)
    safe_set_cell_value(ws, total_row, 8, total_sick)
    safe_set_cell_value(ws, total_row, 9, total_on_duty_count)
    safe_set_cell_value(ws, total_row, 10, total_after_duty_count)
    safe_set_cell_value(ws, total_row, 11, total_training)
    safe_set_cell_value(ws, total_row, 12, total_seconded_from)
    safe_set_cell_value(ws, total_row, 13, total_seconded_to)

    # Название файла
    current_date = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"расход_{department.name}_{current_date}.xlsx"

    # Сохраняем в BytesIO
    file_buffer = BytesIO()
    wb.save(file_buffer)
    file_buffer.seek(0)

    return file_buffer, filename
