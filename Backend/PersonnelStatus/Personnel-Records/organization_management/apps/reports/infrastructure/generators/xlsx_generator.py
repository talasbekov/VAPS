from io import BytesIO
from typing import Tuple

import openpyxl
from openpyxl.styles import Font, Alignment

from organization_management.apps.reports.infrastructure import report_table


class XLSXGenerator:
    """
    Простейший XLSX‑генератор: выводит сводную таблицу по подразделениям.
    Возвращает кортеж (filename, bytes).
    """

    def generate(self, data, report) -> Tuple[str, bytes]:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отчет"

        ws["A1"].value = f"Отчет: {report.get_report_type_display()}"
        ws["A2"].value = f"Раздел: {data.get('division')}"
        ws["A3"].value = f"Дата: {data.get('date')}"
        ws["A1"].font = Font(bold=True)
        ws["A2"].font = Font(italic=True)

        ws.append(report_table.headers(data))
        for cell in ws[4]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row in data.get("rows", []):
            ws.append(report_table.cells(data, row))

        stream = BytesIO()
        wb.save(stream)
        content = stream.getvalue()
        filename = f"report_{report.id}.xlsx"
        return filename, content
