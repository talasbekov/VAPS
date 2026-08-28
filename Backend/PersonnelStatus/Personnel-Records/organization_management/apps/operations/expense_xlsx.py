"""Рендерер .xlsx расхода (порт apps/documents/generators/expense_xlsx.py из
Backend/VAPS).

XLSX — ДОКУМЕНТНАЯ форма, в отличие от машиночитаемого .csv: статусные ячейки
несут не только число, но и поимённый состав. Это разные потребители одной
выгрузки — таблица для дальнейшего счёта и лист для подписи, — и объединять их
в один формат значило бы не годиться ни одному.

Числа остаются ЧИСЛАМИ: безсоставная ячейка пишется целым, а не строкой,
иначе Excel не сложит колонку и сортировка пойдёт лексикографической. Ячейка с
составом — текст «число, затем строки людей»: рич-текста у формата нет, один
шрифт на ячейку, и уместить два размера (как в .docx) невозможно.

Формул в файле нет: все значения предвычислены билдером. Формула, вписанная в
документ, пересчиталась бы у получателя от его же правок — и подпись стояла бы
под числами, которых никто не утверждал.

openpyxl импортируется ЛЕНИВО: раздел не должен требовать библиотеку выгрузки
для сдачи дня и чтения расхода.
"""
from organization_management.apps.operations.expense_layout import (
    event_cell,
    FIXED_HEAD,
    FONT_NAME,
    TABLE_SIZE_PT,
    TITLE_SIZE_PT,
    TOTALS_LABEL,
    document_title,
    header_row,
    member_lines,
)


def _cell_value(cell):
    """Значение статусной ячейки: целое без состава, текст с составом."""
    if not cell.members:
        return cell.count
    return "\n".join([str(cell.count), *member_lines(cell.members)])


def generate_expense_xlsx(data):
    """Данные документа → байты .xlsx."""
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    # Имя листа — ISO-дата: она свободна от запрещённых Excel символов
    # (/\?*[]:) и не зависит от локали читателя. Подпись «для глаз» стоит в
    # титуле, где ей и место, — там она в привычном ДД.ММ.ГГГГ.
    sheet.title = data.business_date.isoformat()

    head = header_row(data.columns)
    width = len(head)

    sheet.cell(row=1, column=1, value=document_title(data))
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    sheet.cell(row=1, column=1).font = Font(
        name=FONT_NAME, size=TITLE_SIZE_PT, bold=True
    )
    sheet.cell(row=1, column=1).alignment = Alignment(
        horizontal="center", vertical="center"
    )

    for index, label in enumerate(head, start=1):
        cell = sheet.cell(row=2, column=index, value=label)
        cell.font = Font(name=FONT_NAME, size=TABLE_SIZE_PT, bold=True)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    row_index = 3
    for number, row in enumerate(data.rows, start=1):
        values = [number, row.name, row.staff_total, row.list_total, row.vacancies]
        values.extend(_cell_value(row.cells[column]) for column in data.columns)
        # Приданные — «+N»: это прибавка СВЕРХ списка, и голое число в ряду
        # колонок списка читалось бы как ещё одна его часть.
        values.append(f"+{row.attached.count}")
        values.append(event_cell(row.event))
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=index, value=value)
            cell.font = Font(name=FONT_NAME, size=TABLE_SIZE_PT)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row_index += 1

    totals = data.totals
    values = ["", TOTALS_LABEL, totals.staff_total, totals.list_total, totals.vacancies]
    values.extend(totals.columns[column] for column in data.columns)
    values.append(f"+{totals.attached}")
    values.append(event_cell(totals.event))
    for index, value in enumerate(values, start=1):
        cell = sheet.cell(row=row_index, column=index, value=value)
        cell.font = Font(name=FONT_NAME, size=TABLE_SIZE_PT, bold=True)

    sheet.column_dimensions[get_column_letter(2)].width = 32
    for index in range(len(FIXED_HEAD) + 1, width + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 24

    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
