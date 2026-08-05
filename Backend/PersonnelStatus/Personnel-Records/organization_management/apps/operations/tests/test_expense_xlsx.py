"""Выгрузка .xlsx: числа числами, состав поимённо, ничего лишнего.

Проверяется ОТКРЫТАЯ книга, а не байты: важен не формат файла (его держит
openpyxl), а то, что получатель увидит и сможет сложить.
"""
import io
from datetime import date

import pytest
from openpyxl import load_workbook

from organization_management.apps.operations.expense_document import (
    build_expense_document,
)
from organization_management.apps.operations.expense_layout import (
    CELL_MAX_MEMBERS,
    COLUMN_LABELS,
    TOTALS_LABEL,
)
from organization_management.apps.operations.expense_xlsx import generate_expense_xlsx
from organization_management.apps.operations.strength_report import StatusCatalog

DAY = date(2026, 8, 4)


@pytest.fixture
def catalog():
    return StatusCatalog.from_rows(
        [
            {
                "code": "IN_SERVICE",
                "priority": 999,
                "report_column_code": "IN_SERVICE",
                "counts_in_staff": True,
            },
            {
                "code": "DUTY",
                "priority": 10,
                "report_column_code": "ON_DUTY",
                "counts_in_staff": True,
            },
            {
                "code": "VACATION",
                "priority": 30,
                "report_column_code": "VACATION",
                "counts_in_staff": True,
            },
        ]
    )


def member(employee_id, full_name="Иванов Иван", rank="капитан"):
    return {"employee_id": employee_id, "full_name": full_name, "rank": rank}


def fact(employee_id, code):
    return {
        "employee_id": employee_id,
        "status_type_code": code,
        "date_start": DAY.isoformat(),
        "date_end": date(2026, 8, 10).isoformat(),
        "source": "USER",
    }


def document(catalog, roster, rows, **overrides):
    kwargs = {
        "catalog": catalog,
        "division_title": "Управление кадров",
        "staff_total": 10,
        "vacancies": 7,
        "attached": 2,
    }
    kwargs.update(overrides)
    return build_expense_document({"roster": roster, "rows": rows}, DAY, **kwargs)


@pytest.fixture
def data(catalog):
    return document(
        catalog,
        [member(1), member(2, full_name="Петров Пётр", rank=""), member(3)],
        [fact(1, "DUTY"), fact(2, "VACATION")],
    )


def sheet_of(data):
    return load_workbook(io.BytesIO(generate_expense_xlsx(data))).active


def column_index(sheet, label):
    for cell in sheet[2]:
        if cell.value == label:
            return cell.column
    raise AssertionError(f"колонка {label!r} не найдена в шапке")


# ── Что видит получатель ─────────────────────────────────────────────────


def test_the_sheet_is_named_by_the_iso_date(data):
    """ISO свободна от запрещённых Excel символов и от локали читателя."""
    assert sheet_of(data).title == "2026-08-04"


def test_the_title_row_spans_the_table(data):
    sheet = sheet_of(data)

    assert "Управление кадров" in sheet.cell(row=1, column=1).value
    assert "04.08.2026" in sheet.cell(row=1, column=1).value
    (merged,) = sheet.merged_cells.ranges
    assert merged.min_row == merged.max_row == 1
    assert merged.max_col == len(list(sheet[2]))


def test_the_header_follows_the_catalog(data):
    header = [cell.value for cell in sheet_of(data)[2]]

    assert header[:5] == ["№", "Управление", "По штату", "По списку", "Вакансии"]
    assert header[5:] == [
        COLUMN_LABELS["ON_DUTY"],
        COLUMN_LABELS["VACATION"],
        COLUMN_LABELS["IN_SERVICE"],
        "Придано (+N)",
    ]


# ── Числа остаются числами ───────────────────────────────────────────────


def test_a_cell_without_members_is_a_number(data):
    """Иначе Excel не сложит колонку, а сортировка пойдёт по буквам."""
    sheet = sheet_of(data)
    cell = sheet.cell(row=3, column=column_index(sheet, COLUMN_LABELS["IN_SERVICE"]))

    assert cell.value == 1
    assert isinstance(cell.value, int)


def test_the_denominator_cells_are_numbers(data):
    sheet = sheet_of(data)

    assert [sheet.cell(row=3, column=index).value for index in (3, 4, 5)] == [10, 3, 7]


def test_the_totals_are_numbers_too(data):
    sheet = sheet_of(data)
    last = sheet.max_row

    assert sheet.cell(row=last, column=2).value == TOTALS_LABEL
    assert sheet.cell(row=last, column=3).value == 10
    assert sheet.cell(row=last, column=1).value in (None, "")


def test_the_file_carries_no_formulas(data):
    """Формула пересчиталась бы у получателя от его же правок.

    Подпись стояла бы под числами, которых никто не утверждал.
    """
    sheet = sheet_of(data)

    for row in sheet.iter_rows():
        for cell in row:
            assert not (isinstance(cell.value, str) and cell.value.startswith("="))


# ── Поимённый состав ─────────────────────────────────────────────────────


def test_a_cell_with_members_shows_the_count_and_the_names(data):
    sheet = sheet_of(data)
    cell = sheet.cell(row=3, column=column_index(sheet, COLUMN_LABELS["ON_DUTY"]))

    lines = cell.value.split("\n")
    assert lines[0] == "1"
    assert lines[1] == "капитан Иванов Иван — 04.08.2026–10.08.2026"


def test_a_member_without_a_rank_has_no_leading_space(data):
    sheet = sheet_of(data)
    cell = sheet.cell(row=3, column=column_index(sheet, COLUMN_LABELS["VACATION"]))

    assert cell.value.split("\n")[1].startswith("Петров Пётр —")


def test_a_long_list_is_truncated_with_an_honest_tail(catalog):
    """Молча обрезанный список выглядел бы как полный.

    Читатель документа считал бы, что видит всех.
    """
    roster = [member(index, full_name=f"Сотрудник {index}") for index in range(30)]
    rows = [fact(index, "DUTY") for index in range(30)]
    sheet = sheet_of(document(catalog, roster, rows))
    cell = sheet.cell(row=3, column=column_index(sheet, COLUMN_LABELS["ON_DUTY"]))

    lines = cell.value.split("\n")
    assert lines[0] == "30"
    assert len(lines) == CELL_MAX_MEMBERS + 2  # число + имена + хвост
    assert lines[-1] == f"… ещё {30 - CELL_MAX_MEMBERS}"


def test_attached_is_shown_as_a_surplus(data):
    """«+N» — прибавка СВЕРХ списка.

    Голое число в ряду колонок списка читалось бы как ещё одна его часть.
    """
    sheet = sheet_of(data)

    assert sheet.cell(row=3, column=column_index(sheet, "Придано (+N)")).value == "+2"
    assert (
        sheet.cell(row=sheet.max_row, column=column_index(sheet, "Придано (+N)")).value
        == "+2"
    )
