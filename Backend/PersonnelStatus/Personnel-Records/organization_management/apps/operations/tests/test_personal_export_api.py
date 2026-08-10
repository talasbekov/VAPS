"""GET /api/operations/daily-submissions/{id}/export/ — личная копия файлом.

Зона вьюхи: право и область, выдача ЗАПРОШЕННОЙ версии, кириллическое имя
файла и то, что в журнал попадает актор из аутентификации.
"""
import io

import pytest
from openpyxl import load_workbook
from rest_framework.test import APIClient

from organization_management.apps.divisions.models import Division
from organization_management.apps.operations import audit_service, clock
from organization_management.apps.operations.day_submission_service import (
    amend_day,
    submit_day,
)
from organization_management.apps.operations.models_audit import OpsAuditLog
from organization_management.apps.operations.models_submission import (
    OpsDailySubmission,
)
from organization_management.apps.operations.tests.test_bulk_status_api import (
    client_for,
)
from organization_management.apps.operations.tests.test_day_submission_service import (
    MORNING,
    TODAY,
    in_slot,
)
from organization_management.apps.operations.tests.test_traffic_light import types  # noqa: F401

pytestmark = pytest.mark.django_db

ACTOR = "7"


@pytest.fixture
def division():
    return Division.objects.create(name="Управление кадров")


def url(submission):
    return f"/api/operations/daily-submissions/{submission.pk}/export/"


def reader(name="pe-reader"):
    api, user = client_for(name, "OBSERVER", ["status.view"])
    return api, user


def submit(division):
    with clock.override(MORNING):
        return submit_day(
            division_id=division.id, business_date=TODAY, actor=ACTOR
        )


def amend(division):
    with clock.override(MORNING):
        return amend_day(
            division_id=division.id,
            business_date=TODAY,
            actor=ACTOR,
            reason="ошибка",
            sanction="замечание",
        )


def get(api, submission):
    with clock.override(MORNING):
        return api.get(url(submission))


# ── Файл ─────────────────────────────────────────────────────────────────


def test_the_response_is_a_workbook(types, division):
    in_slot(division)
    submission = submit(division)
    api, _ = reader()

    response = get(api, submission)

    assert response.status_code == 200
    sheet = load_workbook(io.BytesIO(response.content)).active
    assert sheet.title == TODAY.isoformat()


def test_the_cyrillic_filename_survives_the_header(types, division):
    """Голый filename доехал бы до браузера искажённым.

    Имя копии кириллическое по построению («сдача_…»), и без filename* с
    UTF-8 пользователь получил бы файл со сломанным именем.
    """
    submission = submit(division)
    api, _ = reader()

    response = get(api, submission)

    assert response["Content-Disposition"].startswith("attachment; filename*=UTF-8''")
    assert "%D1%81%D0%B4%D0%B0%D1%87%D0%B0" in response["Content-Disposition"]


def test_the_requested_version_is_exported_not_the_current_one(types, division):
    """Доказывают обычно то заявление, которое потом поправили."""
    in_slot(division)
    first = submit(division)
    amend(division)
    api, _ = reader()

    response = get(api, first)

    sheet = load_workbook(io.BytesIO(response.content)).active
    passport = {row[0].value: row[1].value for row in sheet.iter_rows(max_row=9, max_col=2)}
    assert passport["Версия"] == 1
    assert passport["Действующая"] == "Нет"


# ── Журнал ───────────────────────────────────────────────────────────────


def test_the_log_records_who_took_the_copy(types, division):
    submission = submit(division)
    api, user = reader()

    get(api, submission)

    entry = OpsAuditLog.objects.get(action=audit_service.SUBMISSION_EXPORTED)
    # В журнале должен стоять тот, кто унёс копию, а не имя из запроса.
    assert entry.actor_user_id == str(user.pk)
    assert entry.entity_id == submission.pk


# ── Гарды ────────────────────────────────────────────────────────────────


def test_anonymous_403(types, division):
    submission = submit(division)

    response = get(APIClient(), submission)

    assert response.status_code == 403
    assert not OpsAuditLog.objects.filter(
        action=audit_service.SUBMISSION_EXPORTED
    ).exists()


def test_a_foreign_division_is_403(types, division):
    submission = submit(division)
    foreign = Division.objects.create(name="Чужое")
    api, _ = client_for(
        "pe-scoped", "OPERATOR", ["status.view"], scope_division_id=foreign.id
    )

    response = get(api, submission)

    assert response.status_code == 403
    assert not OpsAuditLog.objects.filter(
        action=audit_service.SUBMISSION_EXPORTED
    ).exists()


def test_the_write_right_alone_does_not_open_the_copy(types, division):
    submission = submit(division)
    api, _ = client_for("pe-writer", "WRITER", ["daily_report.mark_update"])

    assert get(api, submission).status_code == 403


def test_an_unknown_version_is_404(types, division):
    submission = submit(division)
    submission.pk += 10_000

    assert get(reader()[0], submission).status_code == 404


def test_an_unsupported_snapshot_schema_is_422(types, division):
    submission = submit(division)
    OpsDailySubmission.objects.filter(pk=submission.pk).update(
        snapshot={"schema_version": 99, "roster": [], "rows": []}
    )
    api, _ = reader()

    response = get(api, submission)

    assert response.status_code == 422
    assert response.data["error_code"] == "SNAPSHOT_SCHEMA_UNSUPPORTED"
    # Отказ следа не оставляет: журнал означает «файл отдан».
    assert not OpsAuditLog.objects.filter(
        action=audit_service.SUBMISSION_EXPORTED
    ).exists()
