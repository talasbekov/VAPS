"""GET /api/operations/daily-summaries/export/ — сводный расход файлом.

Зона вьюхи: право, порядок гардов, формат и то, что в файл попадают строки
ВСЕХ подразделений сводки. Содержимое строк покрыто test_summary_document.py.
"""
import io

import pytest
from docx import Document as DocxDocument
from openpyxl import load_workbook
from rest_framework.test import APIClient

from organization_management.apps.divisions.models import Division
from organization_management.apps.operations import clock
from organization_management.apps.operations.day_submission_service import submit_day
from organization_management.apps.operations.summary_service import assemble_summary
from organization_management.apps.operations.tests.test_bulk_status_api import (
    client_for,
)
from organization_management.apps.operations.tests.test_day_submission_service import (
    MORNING,
    TODAY,
    in_slot,
)
from organization_management.apps.operations.tests.test_traffic_light import types  # noqa: F401

pytestmark = pytest.mark.django_db

URL = "/api/operations/daily-summaries/export/"
ACTOR = "7"


@pytest.fixture
def tree():
    root = Division.objects.create(name="Управление")
    left = Division.objects.create(name="Первый отдел", parent=root)
    right = Division.objects.create(name="Второй отдел", parent=root)
    in_slot(root)
    in_slot(left)
    in_slot(right)
    return root, left, right


def reader(name="sx-reader"):
    api, _ = client_for(name, "OBSERVER", ["status.view"])
    return api


def get(api, division_id=None, **params):
    if division_id is not None:
        params.setdefault("division_id", division_id)
    with clock.override(MORNING):
        return api.get(URL, params)


def assembled(tree_fixture):
    root, left, right = tree_fixture
    with clock.override(MORNING):
        submit_day(division_id=left.id, business_date=TODAY, actor=ACTOR)
        submit_day(division_id=right.id, business_date=TODAY, actor=ACTOR)
        return assemble_summary(
            division_id=root.id, business_date=TODAY, actor=ACTOR
        )


def csv_rows(response):
    text = response.content.decode("utf-8-sig")
    return [line.split(";") for line in text.split("\r\n") if line]


# ── Файл ─────────────────────────────────────────────────────────────────


def test_the_file_carries_a_row_per_division(types, tree):
    root, _, _ = tree
    assembled(tree)

    rows = csv_rows(get(reader(), root.id))

    # титул + шапка + три строки + ИТОГО
    assert len(rows) == 6
    assert [row[1] for row in rows[2:]] == [
        "Управление",
        "Первый отдел",
        "Второй отдел",
        "ИТОГО",
    ]


def test_the_rows_are_numbered_and_the_total_is_not(types, tree):
    root, _, _ = tree
    assembled(tree)

    rows = csv_rows(get(reader(), root.id))

    assert [row[0] for row in rows[2:]] == ["1", "2", "3", ""]


def test_xlsx_is_a_real_workbook(types, tree):
    root, _, _ = tree
    assembled(tree)

    response = get(reader(), root.id, file_format="xlsx")

    sheet = load_workbook(io.BytesIO(response.content)).active
    assert sheet.title == TODAY.isoformat()
    assert sheet.max_row == 6


def test_docx_carries_every_division_row(types, tree):
    """Многострочный документ печатной формой: строка на подразделение.

    Многострочность — единственное, чем сводка отличается от личной
    выгрузки, и проверять её надо в том формате, где строки и печатаются.
    """
    root, _, _ = tree
    assembled(tree)

    response = get(reader(), root.id, file_format="docx")

    assert response.status_code == 200
    (table,) = DocxDocument(io.BytesIO(response.content)).tables
    assert len(table.rows) == 5  # шапка + три подразделения + ИТОГО
    assert table.rows[-1].cells[1].text == "ИТОГО"


def test_the_file_is_an_attachment(types, tree):
    root, _, _ = tree
    assembled(tree)

    response = get(reader(), root.id)

    assert response["Content-Disposition"].startswith("attachment;")


def test_an_unknown_format_is_400(types, tree):
    root, _, _ = tree
    assembled(tree)

    assert get(reader(), root.id, file_format="pdf").status_code == 400


# ── Гарды ────────────────────────────────────────────────────────────────


def test_anonymous_403(types, tree):
    root, _, _ = tree

    assert get(APIClient(), root.id).status_code == 403


def test_reading_needs_no_right_to_assemble(types, tree):
    """Скачать сводку и собрать её — разные полномочия."""
    root, _, _ = tree
    assembled(tree)

    assert get(reader(), root.id).status_code == 200


def test_a_day_without_a_summary_is_404(types, tree):
    root, _, _ = tree

    response = get(reader(), root.id)

    assert response.status_code == 404
    assert response.data["error_code"] == "DAY_NOT_SUBMITTED"


def test_a_plain_submission_is_400(types, tree):
    _, left, _ = tree
    with clock.override(MORNING):
        submit_day(division_id=left.id, business_date=TODAY, actor=ACTOR)

    assert get(reader(), left.id).status_code == 400


def test_division_id_is_required(types, tree):
    assert get(reader()).status_code == 400


def test_a_foreign_division_is_403_not_404(types, tree):
    root, left, _ = tree
    api, _ = client_for(
        "sx-scoped", "OPERATOR", ["status.view"], scope_division_id=left.id
    )

    assert get(api, root.id).status_code == 403


def test_a_stranger_gets_403_even_for_a_nonexistent_id(types, tree):
    """Порядок гардов: область РАНЬШЕ существования.

    Обратный порядок отвечал бы чужаку 404 на несуществующий id и 403 на
    существующий — то есть 404 стал бы оракулом существования подразделений.
    Проверить это можно только id, который И вне области, И не существует:
    на существующем оба порядка дают 403.
    """
    root, left, _ = tree
    api, _ = client_for(
        "sx-oracle", "OPERATOR", ["status.view"], scope_division_id=left.id
    )

    assert get(api, root.id + 10_000).status_code == 403


def test_an_unknown_division_is_404(types, tree):
    root, _, _ = tree

    assert get(reader(), root.id + 10_000).status_code == 404
