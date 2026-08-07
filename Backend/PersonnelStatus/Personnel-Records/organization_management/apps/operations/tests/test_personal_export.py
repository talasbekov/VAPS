"""Личная копия сданного дня: зеркало снимка и след в журнале.

Файл берут, чтобы предъявлять его в споре, поэтому проверяется ровно то, на
что в споре сошлются: паспорт версии (в том числе «действующая ли она»),
поимённая ведомость без домыслов и запись о самой выдаче.
"""
import io
from datetime import date, datetime, timezone

import pytest
from openpyxl import load_workbook

from organization_management.apps.divisions.models import Division
from organization_management.apps.operations import audit_service, clock
from organization_management.apps.operations.day_submission_service import (
    amend_day,
    submit_day,
)
from organization_management.apps.operations.exceptions import DomainError
from organization_management.apps.operations.models import StatusType
from organization_management.apps.operations.models_audit import OpsAuditLog
from organization_management.apps.operations.models_submission import (
    OpsDailySubmission,
)
from organization_management.apps.operations.personal_export import (
    EMPTY_STATUS_LEGEND,
    ROSTER_TOTAL_LABEL,
    TABLE_COLUMNS,
    build_personal_export_xlsx,
)
from organization_management.apps.operations.personal_export_service import (
    export_submission,
)
from organization_management.apps.operations.tests.test_day_submission_service import (
    MORNING,
    TODAY,
    fact,
    in_slot,
)
from organization_management.apps.operations.tests.test_status_service import seed_types

pytestmark = pytest.mark.django_db

ACTOR = "7"
DAY = date(2026, 8, 4)


# ── Чистый билдер (без БД) ───────────────────────────────────────────────


def workbook_of(**overrides):
    kwargs = {
        "snapshot": {"schema_version": 1, "roster": [], "rows": []},
        "division_title": "Управление кадров",
        "business_date": DAY,
        "version": 1,
        "is_current": True,
        "event_label": "С изменениями",
        "submitted_by": "7",
        "submitted_at_label": "04.08.2026 09:00",
        "late": False,
        "status_names": {"DUTY": "Дежурство"},
    }
    kwargs.update(overrides)
    return load_workbook(io.BytesIO(build_personal_export_xlsx(**kwargs))).active


def passport_of(sheet):
    values = {}
    for row in sheet.iter_rows(min_row=1, max_row=9, max_col=2):
        values[row[0].value] = row[1].value
    return values


def table_of(sheet):
    header_row = next(
        row[0].row for row in sheet.iter_rows(max_col=1) if row[0].value == "№"
    )
    header = [cell.value for cell in sheet[header_row]][: len(TABLE_COLUMNS)]
    body = []
    for row in sheet.iter_rows(min_row=header_row + 1, max_col=len(TABLE_COLUMNS)):
        if row[0].value is None:
            break
        body.append([cell.value for cell in row])
    return header, body


def test_the_passport_names_the_version_and_its_state():
    sheet = workbook_of(version=3, is_current=False, late=True)

    passport = passport_of(sheet)
    assert passport["Версия"] == 3
    # «Действующая ли» — главный вопрос щита: копия доказывает, ЧТО сдано, и
    # вытесненная поправкой версия обязана сама об этом сообщить.
    assert passport["Действующая"] == "Нет"
    assert passport["Опоздание"] == "Да"
    assert passport["Дата"] == "04.08.2026"


def test_a_person_without_statuses_still_takes_a_row():
    """Знаменатель обязан быть виден.

    Умолчание «в строю» сюда не подставляется: обещан снимок, а не вывод.
    """
    sheet = workbook_of(
        snapshot={
            "schema_version": 1,
            "roster": [{"employee_id": 1, "full_name": "Иванов", "rank": "капитан"}],
            "rows": [],
        }
    )

    _, body = table_of(sheet)
    assert len(body) == 1
    assert body[0][:3] == [1, "капитан", "Иванов"]
    assert body[0][3:] == [None, None, None, None, None]


def test_two_statuses_give_two_rows():
    sheet = workbook_of(
        snapshot={
            "schema_version": 1,
            "roster": [{"employee_id": 1, "full_name": "Иванов", "rank": ""}],
            "rows": [
                {
                    "employee_id": 1,
                    "status_type_code": "DUTY",
                    "date_start": "2026-08-04",
                    "date_end": "2026-08-06",
                    "source": "USER",
                },
                {
                    "employee_id": 1,
                    "status_type_code": "DUTY",
                    "date_start": "2026-08-10",
                    "date_end": "2026-08-12",
                    "source": "OM_AUTO",
                },
            ],
        }
    )

    _, body = table_of(sheet)
    assert len(body) == 2
    assert body[0][3:] == ["Дежурство", "DUTY", "04.08.2026", "06.08.2026", "Оператор"]
    assert body[1][-1] == "Автоматически (дежурства)"


def test_a_status_row_for_someone_outside_the_roster_is_still_printed():
    """Расхождение снимка — это как раз то, ради чего копию и берут.

    Проглотить такую строку значило бы спрятать то, что доказывают файлом.
    """
    sheet = workbook_of(
        snapshot={
            "schema_version": 1,
            "roster": [],
            "rows": [
                {
                    "employee_id": 99,
                    "status_type_code": "DUTY",
                    "date_start": "2026-08-04",
                    "date_end": "2026-08-06",
                    "source": "USER",
                }
            ],
        }
    )

    _, body = table_of(sheet)
    assert len(body) == 1
    assert body[0][1:3] == [None, None]  # ФИО и звания нет — но строка есть
    assert body[0][4] == "DUTY"


def test_an_unknown_code_is_printed_as_is():
    """Словарь вырастет позже кода — файл обязан это пережить."""
    sheet = workbook_of(
        snapshot={
            "schema_version": 1,
            "roster": [{"employee_id": 1, "full_name": "Иванов", "rank": ""}],
            "rows": [
                {
                    "employee_id": 1,
                    "status_type_code": "NEW_CODE",
                    "date_start": "не дата",
                    "date_end": "2026-08-06",
                    "source": "НЕЧТО",
                }
            ],
        },
        status_names={},
    )

    _, body = table_of(sheet)
    assert body[0][3:] == ["NEW_CODE", "NEW_CODE", "не дата", "06.08.2026", "НЕЧТО"]


def test_the_footer_reports_the_roster_size():
    sheet = workbook_of(
        snapshot={
            "schema_version": 1,
            "roster": [
                {"employee_id": 1, "full_name": "Иванов", "rank": ""},
                {"employee_id": 2, "full_name": "Петров", "rank": ""},
            ],
            "rows": [],
        }
    )

    texts = [
        cell.value
        for row in sheet.iter_rows(max_col=1)
        if (cell := row[0]).value is not None
    ]
    assert f"{ROSTER_TOTAL_LABEL}: 2" in texts
    assert EMPTY_STATUS_LEGEND in texts


def test_two_copies_of_one_version_come_out_the_same():
    """То, что этот тест охранял с самого начала: копию берут, чтобы предъявить
    её в споре, и две выдачи одной версии обязаны совпасть.

    Раньше это выражалось через порядок снимка («своей сортировки нет») и
    проверялось буквально: пятый перед первым. Порядок с тех пор стал каноном
    раздела — тем же, по которому печатается расход (см.
    test_personal_export_order.py), — но требование к самой копии не изменилось,
    и здесь проверяется именно оно: НЕЗАВИСИМОСТЬ ОТ ПОРЯДКА ВХОДА.

    Вход переставлен намеренно: совпади он с выходом, тест не отличал бы
    «порядок задан правилом» от «повезло».
    """
    people = [
        {"employee_id": 5, "full_name": "Пятый", "rank": ""},
        {"employee_id": 1, "full_name": "Первый", "rank": ""},
    ]

    straight = table_of(
        workbook_of(snapshot={"schema_version": 1, "roster": people, "rows": []})
    )[1]
    swapped = table_of(
        workbook_of(
            snapshot={
                "schema_version": 1,
                "roster": list(reversed(people)),
                "rows": [],
            }
        )
    )[1]

    assert [row[2] for row in straight] == [row[2] for row in swapped]
    assert [row[2] for row in straight] == ["Первый", "Пятый"]


# ── Сервис ───────────────────────────────────────────────────────────────


@pytest.fixture
def types():
    seed_types()


@pytest.fixture
def division():
    return Division.objects.create(name="Управление кадров")


def submit(division):
    with clock.override(MORNING):
        return submit_day(
            division_id=division.id, business_date=TODAY, actor=ACTOR
        )


def test_the_service_fills_what_the_snapshot_lacks(types, division):
    employee = in_slot(division)
    fact(employee, code="DUTY")
    StatusType.objects.filter(code="DUTY").update(name="Дежурство")
    submission = submit(division)

    payload, filename = export_submission(submission=submission, actor=ACTOR)

    sheet = load_workbook(io.BytesIO(payload)).active
    passport = passport_of(sheet)
    assert passport["Подразделение"] == "Управление кадров"
    assert filename == f"сдача_{TODAY.isoformat()}_v1.xlsx"
    _, body = table_of(sheet)
    assert body[0][3] == "Дежурство"


def test_the_submission_time_is_local(types, division):
    """Отдать время без зоны значило бы напечатать UTC под видом местного."""
    submission = submit(division)
    OpsDailySubmission.objects.filter(pk=submission.pk).update(
        submitted_at=datetime(2026, 8, 4, 4, 0, tzinfo=timezone.utc)
    )
    submission.refresh_from_db()

    payload, _ = export_submission(submission=submission, actor=ACTOR)

    sheet = load_workbook(io.BytesIO(payload)).active
    assert passport_of(sheet)["Время сдачи"] == "04.08.2026 09:00"


def test_a_superseded_version_can_still_be_exported(types, division):
    """Именно ради этого копия и существует.

    Доказывают обычно то заявление, которое потом поправили.
    """
    in_slot(division)
    first = submit(division)
    with clock.override(MORNING):
        amend_day(
            division_id=division.id,
            business_date=TODAY,
            actor=ACTOR,
            reason="ошибка",
            sanction="замечание",
        )
    first.refresh_from_db()

    payload, filename = export_submission(submission=first, actor=ACTOR)

    sheet = load_workbook(io.BytesIO(payload)).active
    assert passport_of(sheet)["Действующая"] == "Нет"
    assert filename.endswith("_v1.xlsx")


def test_the_export_is_written_to_the_log(types, division):
    """Журнал означает «файл ОТДАН».

    Копию предъявляют в споре, и выдача без следа обесценила бы обе стороны —
    и файл, и журнал.
    """
    in_slot(division)
    submission = submit(division)

    export_submission(submission=submission, actor=ACTOR)

    entry = OpsAuditLog.objects.get(action=audit_service.SUBMISSION_EXPORTED)
    assert entry.entity_type == audit_service.ENTITY_SUBMISSION
    assert entry.entity_id == submission.pk
    assert entry.actor_user_id == ACTOR
    assert entry.new_value["version"] == 1
    assert entry.new_value["roster_size"] == 1
    # Снимок в журнал не тащится: он весит сотни килобайт, а мерой файла
    # достаточно размера списка.
    assert "roster" not in entry.new_value


# Неподдерживаемой обязана оставаться какая-то БУДУЩАЯ версия, иначе проверка
# «отказ до журнала» стала бы про одни только опечатки. Число здесь сдвигается
# при каждом повышении схемы: 3 ушла со срезом 135, 4 — с 141, 5 — с 142,
# 6 — с 144.
@pytest.mark.parametrize("schema_version", [None, "1", 7, True])
def test_an_unsupported_snapshot_schema_is_422_before_the_log(
    types, division, schema_version
):
    """Отказ ДО генерации и ДО журнала.

    Точный тип, а не isinstance: True прошёл бы как единица, и файл собрался
    бы по чужой раскладке, ничего об этом не сказав.

    Версия 2 из этого списка УШЛА: снимок расширился уровнем должности, и
    читатель её понимает. На её место взята 3 — заведомо будущая: пример
    неподдерживаемой раскладки не должен становиться поддерживаемым, иначе
    тест доказывал бы обратное тому, что заявляет (тот же урок, что с «pdf» в
    примере неизвестного формата).
    """
    submission = submit(division)
    OpsDailySubmission.objects.filter(pk=submission.pk).update(
        snapshot={"schema_version": schema_version, "roster": [], "rows": []}
    )
    submission.refresh_from_db()

    with pytest.raises(DomainError) as exc:
        export_submission(submission=submission, actor=ACTOR)

    assert exc.value.code == "SNAPSHOT_SCHEMA_UNSUPPORTED"
    assert exc.value.http_status == 422
    assert not OpsAuditLog.objects.filter(
        action=audit_service.SUBMISSION_EXPORTED
    ).exists()


def test_a_deleted_division_does_not_stop_the_export(types, division):
    """Доказать сданное нужнее всего как раз про расформированное.

    Со схемы 5 такая копия ещё и НАЗЫВАЕТ подразделение: имя заморожено в
    снимке в момент сдачи, и расформирование его не уносит. Раньше здесь
    печатался голый id — копия выходила, но о ком она, читатель узнавал бы по
    числу.
    """
    submission = submit(division)
    title = division.name
    Division.objects.filter(pk=division.id).delete()

    payload, _ = export_submission(submission=submission, actor=ACTOR)

    sheet = load_workbook(io.BytesIO(payload)).active
    assert passport_of(sheet)["Подразделение"] == title


def test_a_deleted_division_of_an_old_snapshot_still_prints_its_id(
    types, division  # noqa: F811
):
    """У снимков до схемы 5 имени нет, и подставить его неоткуда — тогда id.

    Без этого теста запасной путь был бы мёртвым кодом: все новые снимки имя
    несут, и ветка не исполнялась бы ни разу.
    """
    submission = submit(division)
    submission.snapshot = {
        key: value
        for key, value in submission.snapshot.items()
        if key != "division_title"
    }
    submission.snapshot["schema_version"] = 4
    submission.save(update_fields=["snapshot"])
    division_id = division.id
    Division.objects.filter(pk=division_id).delete()

    payload, _ = export_submission(submission=submission, actor=ACTOR)

    sheet = load_workbook(io.BytesIO(payload)).active
    assert passport_of(sheet)["Подразделение"] == str(division_id)


# ── Подпись статуса заморожена вместе с днём ─────────────────────────────

# Копию берут, чтобы предъявить её в споре, и две выдачи ОДНОЙ версии обязаны
# совпасть — это записано выше отдельным тестом. Подписи статусов до схемы 4
# брались из СЕГОДНЯШНЕГО словаря, и переименование типа меняло уже выданный
# файл: тот же день, та же версия, другая бумага.


@pytest.fixture
def submitted_day():
    from organization_management.apps.operations.tests.test_traffic_light import (
        types as _types,
    )

    division = Division.objects.create(name="Управление 1")
    seed_types()
    StatusType.objects.filter(code="DUTY").update(name="На дежурстве")
    StatusType.objects.get_or_create(
        code="IN_SERVICE",
        defaults={
            "name": "В строю",
            "priority": 999,
            "report_column_code": "IN_SERVICE",
        },
    )
    del _types
    fact(in_slot(division, last_name="Дежурный"), code="DUTY")
    with clock.override(MORNING):
        submit_day(division_id=division.id, business_date=TODAY, actor=ACTOR)
    return OpsDailySubmission.objects.get(
        division_id=division.id, business_date=TODAY
    )


def status_names_in(submission):
    payload, _ = export_submission(submission=submission, actor=ACTOR)
    _, body = table_of(load_workbook(io.BytesIO(payload)).active)
    return [row[3] for row in body]


def test_renaming_a_type_does_not_change_an_issued_copy(submitted_day):
    """Несущий тест: тот же день, та же версия — та же бумага."""
    before = status_names_in(submitted_day)
    assert before == ["На дежурстве"]

    StatusType.objects.filter(code="DUTY").update(name="Дежурство по части")

    assert status_names_in(submitted_day) == before


def test_the_live_dictionary_really_did_change(submitted_day):
    """Иначе тест выше был бы зелёным и у копии, которая подписи не печатает
    вовсе."""
    StatusType.objects.filter(code="DUTY").update(name="Дежурство по части")

    from organization_management.apps.operations.selectors import StatusTypeSelector

    assert StatusTypeSelector.names_map()["DUTY"] == "Дежурство по части"


def test_an_old_snapshot_without_frozen_names_uses_the_live_ones(submitted_day):
    """Схемы 1–3 подписей не несут, и другого источника для них нет.

    Отказ или голый код здесь потеряли бы читаемость всех дней, сданных до
    этого среза.
    """
    submitted_day.snapshot = {
        key: value
        for key, value in submitted_day.snapshot.items()
        if key != "catalog"
    }
    submitted_day.snapshot["schema_version"] = 2
    submitted_day.save(update_fields=["snapshot"])

    StatusType.objects.filter(code="DUTY").update(name="Дежурство по части")

    assert status_names_in(submitted_day) == ["Дежурство по части"]


def test_a_code_absent_from_the_frozen_catalog_falls_back_to_the_live_name():
    """Живые подписи подмешиваются ПОД замороженные, а не заменяются ими:
    печатать голый код там, где подпись существует, незачем."""
    from organization_management.apps.operations.strength_report import names_of

    snapshot = {"catalog": [{"code": "DUTY", "name": "На дежурстве"}]}

    names = names_of(snapshot, {"DUTY": "Переименовано", "STUDY": "Учёба"})

    assert names == {"DUTY": "На дежурстве", "STUDY": "Учёба"}


def test_renaming_the_division_does_not_change_an_issued_copy(types, division):  # noqa: F811
    """Та же заморозка подписи, что у ФИО, звания и названия статуса.

    Паспорт — то, по чему читатель опознаёт, ЧТО он держит; сменись в нём имя
    подразделения, и две копии одной версии перестали бы совпадать.
    """
    submission = submit(division)
    payload, _ = export_submission(submission=submission, actor=ACTOR)
    before = passport_of(load_workbook(io.BytesIO(payload)).active)["Подразделение"]

    Division.objects.filter(pk=division.id).update(name="Управление имени Другого")

    payload, _ = export_submission(submission=submission, actor=ACTOR)
    after = passport_of(load_workbook(io.BytesIO(payload)).active)["Подразделение"]
    assert after == before == division.name


def test_the_live_name_really_did_change(types, division):  # noqa: F811
    """Иначе тест выше был бы зелёным и у копии, которая имя не печатает."""
    from organization_management.apps.operations.selectors import DivisionTreeSelector

    Division.objects.filter(pk=division.id).update(name="Управление имени Другого")

    assert DivisionTreeSelector.names_map([division.id])[division.id] == (
        "Управление имени Другого"
    )


def test_the_printed_document_names_the_division_the_same_way(types, division):  # noqa: F811
    """Два артефакта из ОДНОГО снимка обязаны назвать подразделение одинаково.

    Копия читала замороженное имя, печатный документ — живое, и после
    переименования пара расходилась. Название документа берётся ИЗ САМОГО
    ДОКУМЕНТА, а не пересчитывается тем же правилом: иначе тест сверял бы
    правило с собой (тот же приём, что в срезе 128).
    """
    from organization_management.apps.operations.expense_release import (
        build_submitted_expense_document,
    )

    submission = submit(division)
    Division.objects.filter(pk=division.id).update(name="Управление имени Другого")

    payload, _ = export_submission(submission=submission, actor=ACTOR)
    copy_title = passport_of(load_workbook(io.BytesIO(payload)).active)[
        "Подразделение"
    ]
    document = build_submitted_expense_document(division.id, TODAY)

    assert document.division_title == copy_title == division.name
