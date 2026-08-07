"""Оба артефакта сданного дня не меняются от правок живого мира.

Срез 133 проверил это для СНИМКА и выведенных из него чисел. Документ — другая
вещь: он склеивает снимок с тем, что в снимке не лежит (справочник, подписи
статусов, название подразделения, знаменатель по штату), и каждая такая склейка
была отдельной утечкой живых данных в подписанный день. Их закрывали срезами
135, 141, 142, 143 и 144 — поодиночке, каждую своим тестом.

Здесь проверяется СОСТАВ этих правок: документ строится дважды, между сборками
живой мир меняется всеми доступными способами разом, и обе сборки обязаны
совпасть. Такой тест ловит не только известные утечки, но и следующую — ту, что
кто-нибудь добавит, не подумав.

Артефакта ДВА, и они разные: печатный расход и личная копия. Первый берёт из
живого мира справочник, шапку и знаменатель; вторая — ещё и подписи статусов,
которых в документе нет вовсе. Проверяй файл только документ — тест на
переименование типа оказался бы ВАКУУМНЫМ (проверено пробой: снятие заморозки
подписей его не краснило).

Сравнение идёт по СЕРИАЛИЗОВАННЫМ артефактам целиком, а не по выбранным полям:
перечисляя поля руками, я закрепил бы ровно то, о чём уже знаю.
"""
import json
from datetime import timedelta
from dataclasses import asdict

import pytest

from organization_management.apps.divisions.models import Division
from organization_management.apps.employees.models import Employee
from organization_management.apps.operations import clock
from organization_management.apps.operations.day_submission_service import submit_day
from organization_management.apps.operations.expense_release import (
    build_submitted_expense_document,
)
from organization_management.apps.operations.models_status import OpsEmployeeStatus
from organization_management.apps.operations.status_types import StatusType
from organization_management.apps.operations.tests.test_day_submission_service import (
    ACTOR,
    MORNING,
    TODAY,
    fact,
    in_slot,
)
from organization_management.apps.operations.tests.test_traffic_light import (
    types,  # noqa: F401 — фикстура pytest
)
from organization_management.apps.staff_unit.models import StaffUnit

pytestmark = pytest.mark.django_db


@pytest.fixture
def division():
    return Division.objects.create(name="Первое управление")


@pytest.fixture
def signed(types, division):  # noqa: F811
    """Сданный день: двое своих, у одного дежурство, плюс вакантный слот."""
    on_duty = in_slot(division, last_name="Дежурный")
    plain = in_slot(division, last_name="Обычный")
    StaffUnit.objects.create(division=division, employee=None, index=90001)
    fact(on_duty, code="DUTY")
    with clock.override(MORNING):
        submit_day(division_id=division.id, business_date=TODAY, actor=ACTOR)
    return division, on_duty, plain


def printed(division):
    """ОБА артефакта дня, приведённые к строке: сравнение побайтное."""
    import io as _io

    from openpyxl import load_workbook

    from organization_management.apps.operations.models_submission import (
        OpsDailySubmission,
    )
    from organization_management.apps.operations.personal_export_service import (
        export_submission,
    )

    document = build_submitted_expense_document(division.id, TODAY)
    submission = OpsDailySubmission.objects.get(
        division_id=division.id, business_date=TODAY, is_current=True
    )
    payload, filename = export_submission(submission=submission, actor=ACTOR)
    # Байты .xlsx сравнивать нельзя — внутри архива лежит время создания.
    # Сравниваются ЗНАЧЕНИЯ ячеек: именно их и читает человек.
    sheet = load_workbook(_io.BytesIO(payload)).active
    copy_values = [[cell.value for cell in row] for row in sheet.iter_rows()]
    return json.dumps(
        [asdict(document), copy_values, filename],
        ensure_ascii=False,
        sort_keys=True,
        default=str,
    )


# ── Каждая правка по отдельности ─────────────────────────────────────────


def test_renaming_a_status_type_changes_nothing(signed):
    division, *_ = signed
    before = printed(division)

    StatusType.objects.filter(code="DUTY").update(name="Дежурство по части")

    assert printed(division) == before


def test_moving_a_status_column_changes_nothing(signed):
    division, *_ = signed
    before = printed(division)

    StatusType.objects.filter(code="DUTY").update(report_column_code="ПЕРЕЕХАЛО")

    assert printed(division) == before


def test_reordering_priorities_changes_nothing(signed):
    division, *_ = signed
    before = printed(division)

    StatusType.objects.filter(code="STUDY").update(priority=1)

    assert printed(division) == before


def test_renaming_the_division_changes_nothing(signed):
    division, *_ = signed
    before = printed(division)

    Division.objects.filter(pk=division.pk).update(name="Управление имени Другого")

    assert printed(division) == before


def test_growing_the_staff_changes_nothing(signed):
    division, *_ = signed
    before = printed(division)

    StaffUnit.objects.create(division=division, employee=None, index=90002)

    assert printed(division) == before


def test_shrinking_the_staff_changes_nothing(signed):
    division, _on_duty, plain = signed
    before = printed(division)

    StaffUnit.objects.filter(employee=plain).delete()

    assert printed(division) == before


def test_editing_the_facts_changes_nothing(signed):
    division, on_duty, plain = signed
    before = printed(division)

    with clock.override(MORNING):
        OpsEmployeeStatus.objects.filter(employee_id=on_duty.id).update(
            cancelled_at=clock.Clock.now(), cancelled_by=ACTOR
        )
    fact(plain, code="VACATION")

    assert printed(division) == before


def test_renaming_a_person_changes_nothing(signed):
    division, on_duty, _plain = signed
    before = printed(division)

    Employee.objects.filter(pk=on_duty.pk).update(last_name="Переименованный")

    assert printed(division) == before


# ── Всё разом ────────────────────────────────────────────────────────────


def test_all_of_it_at_once_changes_nothing(signed):
    """Ради этого теста файл и заведён.

    По одной правке каждая может гаситься другой; вместе они сдвигают
    справочник, подписи, шапку, знаменатель и факты одновременно — и именно
    так выглядит месяц эксплуатации, а не отдельный опыт.
    """
    division, on_duty, plain = signed
    before = printed(division)

    StatusType.objects.filter(code="DUTY").update(
        name="Дежурство по части", report_column_code="ПЕРЕЕХАЛО"
    )
    StatusType.objects.filter(code="STUDY").update(priority=1)
    Division.objects.filter(pk=division.pk).update(name="Управление имени Другого")
    StaffUnit.objects.create(division=division, employee=None, index=90003)
    StaffUnit.objects.filter(employee=plain).delete()
    Employee.objects.filter(pk=on_duty.pk).update(last_name="Переименованный")
    with clock.override(MORNING):
        OpsEmployeeStatus.objects.filter(employee_id=on_duty.id).update(
            cancelled_at=clock.Clock.now(), cancelled_by=ACTOR
        )

    assert printed(division) == before


# ── Проверка самой пробы ─────────────────────────────────────────────────


def test_the_document_is_not_empty_to_begin_with(signed):
    """Все тесты выше зелены и у документа, который вообще ничего не печатает.

    Здесь показывается, что сравнивается непустая бумага: строка есть,
    знаменатель ненулевой, человек стоит в колонке дежурства поимённо.
    """
    division, *_ = signed

    document = build_submitted_expense_document(division.id, TODAY)

    row = document.rows[0]
    assert row.staff_total == 3
    assert (row.list_total, row.vacancies) == (2, 1)
    duty_cell = row.cells["DUTY"]
    assert duty_cell.count == 1
    assert duty_cell.members[0].full_name


def test_the_live_world_really_did_change(signed):
    """И что правки настоящие: живой расход на ту же дату после них другой."""
    from organization_management.apps.operations.strength_report import (
        StrengthReportService,
    )

    division, _on_duty, plain = signed
    with clock.override(MORNING):
        before = StrengthReportService.compute(TODAY, division_ids={division.id})

    StaffUnit.objects.filter(employee=plain).delete()

    with clock.override(MORNING):
        after = StrengthReportService.compute(TODAY, division_ids={division.id})
    assert after.rows[0].staff_total != before.rows[0].staff_total


def test_cancelling_an_attached_leg_does_not_shrink_the_plus_n(signed):
    """«+N» приданных — последнее живое число подписанного дня (схема 7).

    Арифметику документа оно не ломало: приданные стоят СВЕРХ равенства «Штат
    == Список + Вне списка + Вакансии». Но дрейфовало так же тихо, как и всё
    остальное.

    ПРАВКА ВЫБРАНА НЕ ПЕРВАЯ ПОПАВШАЯСЯ. Сперва тут стояло подтверждение
    возврата — и все три пробы остались зелёными: подтверждение вступает в силу
    СО СЛЕДУЮЩЕГО дня, нога живёт до конца сегодняшнего, и живое «+N» за
    сегодня не двигалось вовсе. Тест был вакуумным. Живое число меняет ровно
    то, что смотрит селектор: отмена ноги ATTACHED.

    Прикомандирование заводится ДО сдачи — иначе «+N» и в снимке был бы нулём,
    и тест не отличал бы заморозку от совпадения.
    """
    from organization_management.apps.operations.models_status import Secondment
    from organization_management.apps.operations.models_submission import (
        OpsDailySubmission,
    )
    from organization_management.apps.operations.secondment_service import (
        initiate_secondment,
    )

    division, *_ = signed
    home = Division.objects.create(name="Штатное управление")
    guest = in_slot(home, last_name="Приданный")
    with clock.override(MORNING):
        initiate_secondment(
            guest.id,
            to_division_id=division.id,
            date_start=TODAY,
            date_end=TODAY + timedelta(days=5),
            actor=ACTOR,
        )
        # Сдача пересобирается: «+N» обязан попасть в снимок ненулевым.
        OpsDailySubmission.objects.filter(
            division_id=division.id, business_date=TODAY
        ).delete()
        submit_day(division_id=division.id, business_date=TODAY, actor=ACTOR)

    document = build_submitted_expense_document(division.id, TODAY)
    assert document.rows[0].attached.count == 1
    before = printed(division)

    pair = Secondment.objects.get(employee_id=guest.id)
    with clock.override(MORNING):
        OpsEmployeeStatus.objects.filter(pk=pair.in_status_id).update(
            cancelled_at=clock.Clock.now(), cancelled_by=ACTOR
        )

    assert printed(division) == before


def test_that_cancellation_really_moves_the_live_number(signed):
    """Иначе тест выше был бы зелёным и у правки, которой «+N» не касается —
    как это и вышло с подтверждением возврата в первом наборе."""
    from organization_management.apps.operations.models_status import Secondment
    from organization_management.apps.operations.secondment_service import (
        initiate_secondment,
    )
    from organization_management.apps.operations.selectors import SecondmentSelector

    division, *_ = signed
    home = Division.objects.create(name="Штатное управление")
    guest = in_slot(home, last_name="Приданный")
    with clock.override(MORNING):
        initiate_secondment(
            guest.id,
            to_division_id=division.id,
            date_start=TODAY,
            date_end=TODAY + timedelta(days=5),
            actor=ACTOR,
        )

    live = SecondmentSelector.attached_counts_on(TODAY, division_ids=[division.id])
    assert live.get(division.id) == 1

    pair = Secondment.objects.get(employee_id=guest.id)
    with clock.override(MORNING):
        OpsEmployeeStatus.objects.filter(pk=pair.in_status_id).update(
            cancelled_at=clock.Clock.now(), cancelled_by=ACTOR
        )

    after = SecondmentSelector.attached_counts_on(TODAY, division_ids=[division.id])
    assert after.get(division.id, 0) == 0


def test_a_frozen_zero_stays_zero_when_someone_is_attached_later(signed):
    """Ноль приданных — законное и ЧАСТОЕ состояние, и он тоже заморожен.

    Признаком «есть ли замороженное число» служит НАЛИЧИЕ ключа, а не его
    истинность: проверяй код `or`, и ноль уходил бы на живое число — то есть
    ровно у большинства дней заморозки бы и не было. Проба «признак —
    истинность» краснит именно здесь.
    """
    from organization_management.apps.operations.secondment_service import (
        initiate_secondment,
    )

    division, *_ = signed
    document = build_submitted_expense_document(division.id, TODAY)
    assert document.rows[0].attached.count == 0
    before = printed(division)

    home = Division.objects.create(name="Штатное управление")
    guest = in_slot(home, last_name="Приданный")
    with clock.override(MORNING):
        initiate_secondment(
            guest.id,
            to_division_id=division.id,
            date_start=TODAY,
            date_end=TODAY + timedelta(days=5),
            actor=ACTOR,
        )

    assert printed(division) == before
