"""Личная копия и расход перечисляют людей одинаково.

Копию берут не саму по себе — её кладут РЯДОМ С РАСХОДОМ, чтобы показать, что
именно было сдано. Значит два файла из одного снимка за один день сравнивают
построчно, и разный порядок в этой паре читается как разные данные: объяснять,
что «просто сортировка другая», придётся уже в разбирательстве.

Расход канон применял, копия — нет: она шла в порядке снимка, то есть по
employee_id, в порядке появления строк в базе. Старший по должности оказывался
между рядовыми потому, что его завели позже, — и только в одном из двух файлов.

ФИКСТУРА ПОДОБРАНА ТАК, ЧТОБЫ ПОРЯДКИ РАЗОШЛИСЬ. Совпади канонический порядок с
порядком по employee_id, тест был бы зелёным и до правки — то есть не проверял
бы ничего. Здесь employee_id растёт, а должности идут от младшей к старшей,
так что канон переворачивает список целиком.
"""
from datetime import date

from openpyxl import load_workbook

from organization_management.apps.operations.expense_document import (
    build_expense_document,
)
from organization_management.apps.operations.personal_export import (
    TABLE_COLUMNS,
    build_personal_export_xlsx,
)
from organization_management.apps.operations.roster_order import order_roster
from organization_management.apps.operations.strength_report import StatusCatalog

DAY = date(2026, 8, 4)
NAME_COLUMN = TABLE_COLUMNS.index("ФИО") + 1

# id растёт, уровень должности падает (меньше — старше). Канон обязан
# перевернуть этот список; порядок снимка оставил бы его как есть.
ROSTER = [
    {"employee_id": 1, "full_name": "Яковлев Я.Я.", "rank": "рядовой",
     "position_level": 30},
    {"employee_id": 2, "full_name": "Борисов Б.Б.", "rank": "сержант",
     "position_level": 20},
    {"employee_id": 3, "full_name": "Абрамов А.А.", "rank": "майор",
     "position_level": 10},
]

SNAPSHOT = {"schema_version": 2, "roster": ROSTER, "rows": []}

PASSPORT = dict(
    division_title="Первое управление",
    business_date=DAY,
    version=1,
    is_current=True,
    event_label="Сдача",
    submitted_by="7",
    submitted_at_label="04.08.2026 08:00",
    late=False,
    status_names={},
)


def names_in_export(snapshot):
    """ФИО в том порядке, в каком они попали в лист."""
    import io

    blob = build_personal_export_xlsx(snapshot=snapshot, **PASSPORT)
    sheet = load_workbook(io.BytesIO(blob)).active
    header = next(
        row
        for row in range(1, sheet.max_row + 1)
        if sheet.cell(row=row, column=1).value == TABLE_COLUMNS[0]
    )
    found = []
    for row in range(header + 1, sheet.max_row + 1):
        value = sheet.cell(row=row, column=NAME_COLUMN).value
        if value:
            found.append(value)
    return found


def test_the_copy_lists_people_in_the_canonical_order():
    """Несущий тест: порядок в файле — канон, а не порядок снимка."""
    assert names_in_export(SNAPSHOT) == [
        member["full_name"] for member in order_roster(ROSTER)
    ]


def test_that_order_is_not_the_snapshot_order():
    """Иначе тест выше был бы зелёным и без правки — фикстура обязана
    РАЗЛИЧАТЬ два порядка, а не совпадать с обоими сразу."""
    snapshot_order = [member["full_name"] for member in ROSTER]

    assert names_in_export(SNAPSHOT) != snapshot_order


def test_the_copy_agrees_with_the_printed_expense():
    """То, ради чего правка и сделана: два файла из ОДНОГО снимка.

    Порядок расхода здесь не пересчитывается своей копией канона, а БЕРЁТСЯ ИЗ
    ПОСТРОЕННОГО ДОКУМЕНТА — иначе тест сверял бы канон с самим собой и остался
    бы зелёным, даже если печатная форма перестанет его применять.

    Все трое получают один и тот же статус, чтобы попасть в ОДНУ колонку: люди
    из разных колонок в документе не сравнимы по порядку, а сравнить надо
    именно сквозной список.
    """
    catalog = StatusCatalog.from_rows(
        [
            {"code": "IN_SERVICE", "priority": 999,
             "report_column_code": "IN_SERVICE", "counts_in_staff": True},
            {"code": "DUTY", "priority": 10, "report_column_code": "ON_DUTY",
             "counts_in_staff": True},
        ]
    )
    snapshot = {
        "schema_version": 2,
        "roster": ROSTER,
        "rows": [
            {
                "employee_id": member["employee_id"],
                "status_type_code": "DUTY",
                "status_id": member["employee_id"],
                "date_start": "2026-08-04",
                "date_end": "2026-08-06",
                "source": "USER",
            }
            for member in ROSTER
        ],
    }

    document = build_expense_document(
        snapshot,
        DAY,
        catalog=catalog,
        division_title="Первое управление",
        staff_total=len(ROSTER),
        vacancies=0,
    )
    printed = [
        cell_member.full_name
        for row in document.rows
        for cell in row.cells.values()
        for cell_member in cell.members
    ]
    assert len(printed) == len(ROSTER), "все трое обязаны попасть в одну колонку"

    assert printed == names_in_export(snapshot)


def test_a_snapshot_without_position_levels_still_lists_everyone():
    """Схема 1 уровня должности не несёт.

    Канон ставит такого человека в конец своей группы и сортирует по фамилии —
    падать на старой версии копия не имеет права: её берут именно за старые
    дни.
    """
    old = {
        "schema_version": 1,
        "roster": [
            {k: v for k, v in member.items() if k != "position_level"}
            for member in ROSTER
        ],
        "rows": [],
    }

    assert names_in_export(old) == ["Абрамов А.А.", "Борисов Б.Б.", "Яковлев Я.Я."]


def test_rows_of_an_unknown_employee_stay_in_the_tail():
    """Канон их не касается: человека, по которому сортировать, у них нет.

    Уехали бы они в середину — расхождение снимка, ради которого копию и
    берут, потерялось бы среди обычных строк.
    """
    with_orphan = {
        "schema_version": 2,
        "roster": ROSTER,
        "rows": [
            {
                "employee_id": 999,
                "status_type_code": "DUTY",
                "status_id": 1,
                "date_start": "2026-08-04",
                "date_end": "2026-08-06",
                "source": "USER",
            }
        ],
    }

    import io

    blob = build_personal_export_xlsx(snapshot=with_orphan, **PASSPORT)
    sheet = load_workbook(io.BytesIO(blob)).active
    header = next(
        row
        for row in range(1, sheet.max_row + 1)
        if sheet.cell(row=row, column=1).value == TABLE_COLUMNS[0]
    )
    code_column = TABLE_COLUMNS.index("Код статуса") + 1

    # Строка сироты — последняя из четырёх, и ФИО у неё пусто.
    last = header + 4
    assert sheet.cell(row=last, column=code_column).value == "DUTY"
    assert not sheet.cell(row=last, column=NAME_COLUMN).value
