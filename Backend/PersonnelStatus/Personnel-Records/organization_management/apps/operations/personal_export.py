"""Личная копия сданного дня — «щит» (порт personal_export.py из Backend/VAPS).

Это не расход и не документ учёта, а ЗЕРКАЛО СНИМКА: паспорт версии сдачи и
поимённая ведомость «кто — с каким статусом — по какой период». Ничего не
выводится и не досчитывается: тот, кто скачал файл, обязан увидеть ровно то,
что лежит в снимке, — иначе спорить им будет не о чем.

Ведомость строится по ПАРЕ (человек, статус): два статуса дают две строки,
человек без статусов — одну с пустыми ячейками (знаменатель обязан быть
виден). Строка на сотрудника вне списка тоже печатается — с пустыми ФИО:
такое расхождение снимка это как раз то, ради чего копию и берут, и молча
проглотить его хуже, чем показать.

Порядок — КАНОН РАЗДЕЛА (`roster_order.order_roster`), тот же, по которому
печатается расход.

Раньше здесь стоял порядок снимка, то есть employee_id, и объяснялось это так:
своей сортировки нет, потому что две копии одной версии обязаны совпасть
побайтно, а второй порядок разошёлся бы с первым. Довод верен против СВОЕЙ
сортировки — но канон не второй порядок, а тот же самый: чистая функция от
полей, которые в снимке уже заморожены (ФИО и уровень должности). Побайтного
совпадения двух копий он не трогает.

А вот расхождение убирает настоящее. Расход канон применял, личная копия — нет,
и два файла, собранные из ОДНОГО снимка за ОДИН день, перечисляли людей
по-разному. Копию берут именно затем, чтобы положить её рядом с расходом;
разный порядок в этой паре читается как разные данные, и объяснять его пришлось
бы уже в разбирательстве.

Снимок схемы 1 уровня должности не несёт: канон ставит такого человека в конец
своей группы и дальше сортирует по фамилии. Расход на тех же данных делает
ровно то же самое — пара остаётся согласованной и на старых версиях.

Даты печатаются как есть, полуинтервалом [с, по): «минус день» здесь не
вычитается — файл показывает снимок, а не толкует его. Непарсящееся значение
выводится дословно и не роняет выгрузку.

Чистая функция: ни ORM, ни часов, ни диска. Всё, чего в снимке нет (имя
подразделения, атрибуты версии, подписи типов), приходит готовым.
"""
from datetime import date

from organization_management.apps.operations.roster_order import order_roster

DATE_FORMAT = "%d.%m.%Y"
YES = "Да"
NO = "Нет"

TABLE_COLUMNS = (
    "№",
    "Звание",
    "ФИО",
    "Статус",
    "Код статуса",
    "С",
    "По",
    "Источник",
)

# Подписи источника факта. Незнакомое значение печатается само: словарь
# вырастет позже кода, и файл обязан это пережить, а не упасть.
SOURCE_LABELS = {
    "USER": "Оператор",
    "KU_SYNC": "Синхронизация КУ",
    "OM_AUTO": "Автоматически (дежурства)",
}

EMPTY_STATUS_LEGEND = "Пустая строка статуса — в снимке статуса нет"
ROSTER_TOTAL_LABEL = "Всего в списочном составе"

_COLUMN_WIDTHS = (6, 16, 32, 24, 22, 13, 13, 26)


def _format_iso_date(value):
    """ISO-строка → ДД.ММ.ГГГГ; непарсящееся значение — дословно."""
    try:
        return date.fromisoformat(value).strftime(DATE_FORMAT)
    except (TypeError, ValueError):
        return value


def _pair_rows(roster, rows):
    """[(человек | {}, статус | None)] в каноническом порядке раздела.

    Человек без статусов даёт ровно одну пару; строки на неизвестного
    сотрудника уходят в хвост с пустой записью списка — канон их не касается,
    потому что человека, по которому сортировать, у них нет.
    """
    grouped = {}
    for row in rows:
        grouped.setdefault(row.get("employee_id"), []).append(row)

    paired = []
    for member in order_roster(roster):
        member_rows = grouped.pop(member.get("employee_id"), [])
        if member_rows:
            paired.extend((member, row) for row in member_rows)
        else:
            paired.append((member, None))
    for orphan_rows in grouped.values():
        paired.extend(({}, row) for row in orphan_rows)
    return paired


def _table_values(number, member, row, status_names, source_labels):
    """Значения одной строки ведомости в порядке TABLE_COLUMNS."""
    values = [number, member.get("rank") or "", member.get("full_name") or ""]
    if row is None:
        # Статуса в снимке нет — ячейки пустые. Подставить сюда «в строю»
        # значило бы напечатать ВЫВОД там, где обещан снимок.
        return values + ["", "", "", "", ""]
    code = row.get("status_type_code")
    source = row.get("source")
    return values + [
        status_names.get(code, code),
        code,
        _format_iso_date(row.get("date_start")),
        _format_iso_date(row.get("date_end")),
        source_labels.get(source, source),
    ]


def build_personal_export_xlsx(
    *,
    snapshot,
    division_title,
    business_date,
    version,
    is_current,
    event_label,
    submitted_by,
    submitted_at_label,
    late,
    status_names,
    source_labels=None,
) -> bytes:
    """Снимок сдачи и атрибуты версии → байты .xlsx личной копии.

    `submitted_at_label` — ГОТОВАЯ строка, а не datetime: openpyxl не пишет в
    ячейку значение с зоной, а приводить время к локальному поясу — забота
    того, кто знает часы раздела.
    """
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    labels = SOURCE_LABELS if source_labels is None else source_labels
    bold = Font(bold=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = business_date.isoformat()

    roster = snapshot.get("roster") or []
    rows = snapshot.get("rows") or []

    passport = (
        ("Подразделение", division_title),
        ("Дата", business_date.strftime(DATE_FORMAT)),
        ("Версия", version),
        ("Событие", event_label),
        # Действующая ли это версия — главный вопрос «щита»: копию берут,
        # чтобы доказать, ЧТО именно было сдано, и вытесненная поправкой
        # версия обязана сама об этом сообщать.
        ("Действующая", YES if is_current else NO),
        ("Сдал", submitted_by),
        ("Время сдачи", submitted_at_label),
        ("Опоздание", YES if late else NO),
        ("Версия схемы снимка", snapshot.get("schema_version")),
    )
    for index, (label, value) in enumerate(passport, start=1):
        sheet.cell(row=index, column=1, value=label).font = bold
        sheet.cell(row=index, column=2, value=value)

    header_row = len(passport) + 2
    for column, label in enumerate(TABLE_COLUMNS, start=1):
        sheet.cell(row=header_row, column=column, value=label).font = bold

    paired = _pair_rows(roster, rows)
    for number, (member, row) in enumerate(paired, start=1):
        values = _table_values(number, member, row, status_names, labels)
        for column, value in enumerate(values, start=1):
            sheet.cell(row=header_row + number, column=column, value=value)

    # Пустая строка-разделитель: подвал не должен читаться хвостом таблицы.
    footer_row = header_row + len(paired) + 2
    sheet.cell(row=footer_row, column=1, value=f"{ROSTER_TOTAL_LABEL}: {len(roster)}")
    sheet.cell(row=footer_row + 1, column=1, value=EMPTY_STATUS_LEGEND)

    for column, width in enumerate(_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
