"""Story 20.4d (FR-40): build_audit_xlsx() — .xlsx-зеркало
build_audit_csv() (20.4c), тот же allowlist-запрет old_value/new_value."""

import datetime
import io

from openpyxl import load_workbook

from apps.core.exports.audit_csv import AUDIT_CSV_COLUMNS
from apps.core.exports.audit_xlsx import build_audit_xlsx


def decode_rows(xlsx_bytes):
    sheet = load_workbook(io.BytesIO(xlsx_bytes)).active
    header = [cell.value for cell in sheet[1]]
    rows = []
    for excel_row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(header, excel_row)))
    return rows


def row(**overrides):
    base = {
        "id": "33333333-3333-3333-3333-333333333333",
        "actor_user_id": "operator-1",
        "action": "EMPLOYEE_UPDATED",
        "entity_type": "employee",
        "entity_id": "11111111-1111-1111-1111-111111111111",
        "request_id": "req-abc123",
        "ip_address": "10.0.0.5",
        "created_at": "2026-07-08T09:00:00+00:00",
    }
    base.update(overrides)
    return base


def test_workbook_opens_without_exception():
    result = build_audit_xlsx([])
    workbook = load_workbook(io.BytesIO(result))
    assert workbook.active is not None


def test_header_row_has_exactly_the_metadata_columns_bold():
    result = build_audit_xlsx([])
    sheet = load_workbook(io.BytesIO(result)).active
    header_cells = sheet[1]
    header = [cell.value for cell in header_cells]
    assert header == list(AUDIT_CSV_COLUMNS)
    assert "old_value" not in header
    assert "new_value" not in header
    assert all(cell.font.bold for cell in header_cells)


def test_sheet_has_a_meaningful_title():
    result = build_audit_xlsx([])
    sheet = load_workbook(io.BytesIO(result)).active
    assert sheet.title != "Sheet"


def test_metadata_row_renders_all_fields():
    result = build_audit_xlsx([row()])
    decoded = decode_rows(result)
    assert decoded[0]["actor_user_id"] == "operator-1"
    assert decoded[0]["ip_address"] == "10.0.0.5"


def test_raw_datetime_normalizes_to_isoformat():
    ts = datetime.datetime(2026, 7, 8, 9, 0, tzinfo=datetime.timezone.utc)
    result = build_audit_xlsx([row(created_at=ts)])
    decoded = decode_rows(result)
    assert decoded[0]["created_at"] == ts.isoformat()


def test_empty_rows_gives_header_only():
    result = build_audit_xlsx([])
    sheet = load_workbook(io.BytesIO(result)).active
    assert sheet.max_row == 1


def test_formula_injection_prefix_is_sanitized():
    result = build_audit_xlsx([row(actor_user_id="=cmd|'/c calc'!A1")])
    decoded = decode_rows(result)
    assert decoded[0]["actor_user_id"].startswith("'=")


def test_payload_keys_in_row_do_not_leak_into_xlsx():
    dirty_row = row(
        old_value={"iin": "900101300123", "full_name": "Иванов"},
        new_value={"iin": "900101300124", "full_name": "Иванов И."},
    )
    result = build_audit_xlsx([dirty_row])

    sheet = load_workbook(io.BytesIO(result)).active
    all_values = []
    for excel_row in sheet.iter_rows(values_only=True):
        all_values.extend(str(v) for v in excel_row if v is not None)
    flattened = " ".join(all_values)

    assert "old_value" not in flattened
    assert "new_value" not in flattened
    assert "900101300123" not in flattened
    assert "900101300124" not in flattened
    decoded = decode_rows(result)
    assert set(decoded[0].keys()) == set(AUDIT_CSV_COLUMNS)


def test_columns_never_include_payload_fields():
    assert "old_value" not in AUDIT_CSV_COLUMNS
    assert "new_value" not in AUDIT_CSV_COLUMNS
    assert "reason" not in AUDIT_CSV_COLUMNS
