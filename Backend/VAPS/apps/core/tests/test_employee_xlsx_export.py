"""Story 20.4d (FR-40): build_employees_xlsx() — .xlsx-зеркало
build_employees_csv() (20.4a), та же маскировка/NFR-4-дисциплина."""

import io

import pytest
from django.db import connection
from django.test.utils import CaptureQueriesContext
from openpyxl import load_workbook

from apps.core.exports.employee_csv import EMPLOYEE_CSV_COLUMNS
from apps.core.exports.employee_xlsx import build_employees_xlsx
from apps.core.models import SensitiveFieldPolicy

pytestmark = pytest.mark.django_db


def decode_rows(xlsx_bytes):
    sheet = load_workbook(io.BytesIO(xlsx_bytes)).active
    header = [cell.value for cell in sheet[1]]
    rows = []
    for excel_row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(header, excel_row)))
    return rows


def test_workbook_opens_without_exception():
    result = build_employees_xlsx([], user_permissions=set())
    workbook = load_workbook(io.BytesIO(result))
    assert workbook.active is not None


def test_header_row_has_expected_columns_bold():
    result = build_employees_xlsx([], user_permissions=set())
    sheet = load_workbook(io.BytesIO(result)).active
    header_cells = sheet[1]
    assert [cell.value for cell in header_cells] == list(EMPLOYEE_CSV_COLUMNS)
    assert all(cell.font.bold for cell in header_cells)


def test_sheet_has_a_meaningful_title():
    result = build_employees_xlsx([], user_permissions=set())
    sheet = load_workbook(io.BytesIO(result)).active
    assert sheet.title != "Sheet"


def test_iin_masked_without_permission():
    SensitiveFieldPolicy.objects.create(
        field_code="iin",
        permission_code="employee.sensitive.view",
        mask_strategy="PARTIAL_MASK",
    )
    rows = [{"iin": "900101300123", "full_name": "Иванов"}]

    result = build_employees_xlsx(rows, user_permissions=set())

    decoded = decode_rows(result)
    assert decoded[0]["iin"] != "900101300123"
    assert decoded[0]["iin"].endswith("0123")


def test_iin_visible_with_permission():
    SensitiveFieldPolicy.objects.create(
        field_code="iin",
        permission_code="employee.sensitive.view",
        mask_strategy="PARTIAL_MASK",
    )
    rows = [{"iin": "900101300123", "full_name": "Иванов"}]

    result = build_employees_xlsx(rows, user_permissions={"employee.sensitive.view"})

    decoded = decode_rows(result)
    assert decoded[0]["iin"] == "900101300123"


def test_empty_rows_produce_header_only():
    result = build_employees_xlsx([], user_permissions=set())
    sheet = load_workbook(io.BytesIO(result)).active
    assert sheet.max_row == 1


def test_special_characters_round_trip():
    """Ревью (Edge Case Hunter): портировано из test_employee_csv_export.py
    — единственная строка того файла, изначально пропущенная при переносе
    на xlsx (CSV- и xlsx-экранирование — разные механизмы, покрытие не
    наследуется автоматически)."""
    rows = [{"full_name": 'Иванов, Иван "Ваня"'}]

    result = build_employees_xlsx(rows, user_permissions=set())

    decoded = decode_rows(result)
    assert decoded[0]["full_name"] == 'Иванов, Иван "Ваня"'


@pytest.mark.parametrize("payload", ["=cmd|'/c calc'!A0", "+1+1", "-2+3", "@SUM(1,1)"])
def test_formula_injection_is_neutralized(payload):
    rows = [{"full_name": payload}]

    result = build_employees_xlsx(rows, user_permissions=set())

    decoded = decode_rows(result)
    assert decoded[0]["full_name"] == "'" + payload


@pytest.mark.parametrize("row_count", [0, 1, 10])
def test_single_query_regardless_of_row_count(row_count):
    """Ревью (Blind Hunter): параметризовано по 0/1/N — один прогон на
    N=10 доказывает «ровно 1 запрос на 10 строк», НЕ доказывает
    независимость от числа строк (off-by-one на N=0/1 остался бы
    незамеченным)."""
    rows = [{"full_name": f"Сотрудник {i}"} for i in range(row_count)]

    with CaptureQueriesContext(connection) as ctx:
        build_employees_xlsx(rows, user_permissions=set())

    assert len(ctx.captured_queries) == 1
