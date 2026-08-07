"""Story 20.4d (FR-40): build_history_xlsx() — .xlsx-зеркало
build_history_csv() (20.4b)."""

import datetime
import io

from openpyxl import load_workbook

from apps.core.exports.history_csv import HISTORY_CSV_COLUMNS
from apps.core.exports.history_xlsx import build_history_xlsx


def decode_rows(xlsx_bytes):
    sheet = load_workbook(io.BytesIO(xlsx_bytes)).active
    header = [cell.value for cell in sheet[1]]
    rows = []
    for excel_row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(header, excel_row)))
    return rows


def row(**overrides):
    base = {
        "employee_id": "11111111-1111-1111-1111-111111111111",
        "employee_full_name": "Иванов Иван Иванович",
        "division_id": "22222222-2222-2222-2222-222222222222",
        "division_name": "Отдел А",
        "starts_at": "2026-07-01T00:00:00+00:00",
        "ends_at": "2026-07-31T00:00:00+00:00",
        "source": "MANUAL",
    }
    base.update(overrides)
    return base


def test_workbook_opens_without_exception():
    result = build_history_xlsx([])
    workbook = load_workbook(io.BytesIO(result))
    assert workbook.active is not None


def test_header_row_has_expected_columns_bold():
    result = build_history_xlsx([])
    sheet = load_workbook(io.BytesIO(result)).active
    header_cells = sheet[1]
    assert [cell.value for cell in header_cells] == list(HISTORY_CSV_COLUMNS)
    assert all(cell.font.bold for cell in header_cells)


def test_sheet_has_a_meaningful_title():
    result = build_history_xlsx([])
    sheet = load_workbook(io.BytesIO(result)).active
    assert sheet.title != "Sheet"


def test_closed_transfer_has_both_dates():
    result = build_history_xlsx([row()])
    decoded = decode_rows(result)
    assert decoded[0]["starts_at"] == "2026-07-01T00:00:00+00:00"
    assert decoded[0]["ends_at"] == "2026-07-31T00:00:00+00:00"


def test_open_transfer_renders_empty_ends_at():
    result = build_history_xlsx([row(ends_at=None)])
    decoded = decode_rows(result)
    assert decoded[0]["ends_at"] in ("", None)


def test_empty_rows_gives_header_only():
    result = build_history_xlsx([])
    sheet = load_workbook(io.BytesIO(result)).active
    assert sheet.max_row == 1


def test_raw_datetime_normalizes_to_isoformat():
    ts = datetime.datetime(2026, 7, 1, 12, 30, tzinfo=datetime.timezone.utc)
    result = build_history_xlsx([row(starts_at=ts)])
    decoded = decode_rows(result)
    assert decoded[0]["starts_at"] == ts.isoformat()


def test_formula_injection_prefix_is_sanitized():
    result = build_history_xlsx([row(employee_full_name="=cmd|'/c calc'!A1")])
    decoded = decode_rows(result)
    assert decoded[0]["employee_full_name"].startswith("'=")


def test_comma_and_quote_round_trip():
    tricky = 'Иванов, Иван "Ваня"'
    result = build_history_xlsx([row(employee_full_name=tricky)])
    decoded = decode_rows(result)
    assert decoded[0]["employee_full_name"] == tricky
