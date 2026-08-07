"""Story 20.4d (FR-40): `.xlsx`-зеркало `audit_csv.py` (20.4c) — тот же
allowlist из 8 метаданных-колонок, БЕЗ `old_value`/`new_value`/`reason`
(та же дисциплина, что CSV-версия — колонки строятся ТОЛЬКО из
`AUDIT_CSV_COLUMNS`, не из `row.items()`)."""

import io

from apps.core.exports.audit_csv import AUDIT_CSV_COLUMNS
from apps.core.exports.csv_safety import normalize_value, sanitize_cell


def build_audit_xlsx(rows) -> bytes:
    """`rows` — список УЖЕ загруженных построчных `dict` (та же
    ответственность вызывающего кода, что `build_audit_csv`)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Аудит"

    bold = Font(bold=True)
    for column, name in enumerate(AUDIT_CSV_COLUMNS, start=1):
        sheet.cell(row=1, column=column, value=name).font = bold

    for row_index, row in enumerate(rows, start=2):
        for column, name in enumerate(AUDIT_CSV_COLUMNS, start=1):
            value = sanitize_cell(normalize_value(row.get(name)))
            sheet.cell(row=row_index, column=column, value=value)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
