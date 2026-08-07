"""Story 20.4d (FR-40): `.xlsx`-зеркало `history_csv.py` (20.4b) — тот же
набор колонок, БЕЗ маскирования (нет чувствительных полей в строке
истории), `datetime`-нормализация через общий `normalize_value()`."""

import io

from apps.core.exports.csv_safety import normalize_value, sanitize_cell
from apps.core.exports.history_csv import HISTORY_CSV_COLUMNS


def build_history_xlsx(rows) -> bytes:
    """`rows` — список УЖЕ загруженных построчных `dict` (та же
    ответственность вызывающего кода, что `build_history_csv`)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "История переводов"

    bold = Font(bold=True)
    for column, name in enumerate(HISTORY_CSV_COLUMNS, start=1):
        sheet.cell(row=1, column=column, value=name).font = bold

    for row_index, row in enumerate(rows, start=2):
        for column, name in enumerate(HISTORY_CSV_COLUMNS, start=1):
            value = sanitize_cell(normalize_value(row.get(name)))
            sheet.cell(row=row_index, column=column, value=value)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
