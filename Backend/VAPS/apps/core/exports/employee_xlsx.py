"""Story 20.4d (FR-40): `.xlsx`-зеркало `employee_csv.py` (20.4a) —
буквально та же маскировка (`mask_employee_data()`, ОДИН запрос
`SensitiveFieldPolicy` снаружи цикла, NFR-4), тот же набор колонок,
рендер — `openpyxl` вместо `csv`. `sanitize_cell()` применён ЯВНО (в
отличие от существующих `expense_xlsx.py`/`personal_export.py`, которые
этого не делают — их собственный пробел, не эта стори, CWE-1236
одинаково актуальна для .xlsx, открытого в Excel/LibreOffice)."""

import io

from apps.core.exports.csv_safety import normalize_value, sanitize_cell
from apps.core.exports.employee_csv import EMPLOYEE_CSV_COLUMNS
from apps.core.models import SensitiveFieldPolicy
from apps.core.services import mask_employee_data


def build_employees_xlsx(rows, *, user_permissions) -> bytes:
    """`rows` — список УЖЕ загруженных построчных `dict` (та же
    ответственность вызывающего кода, что `build_employees_csv`)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    policies = list(SensitiveFieldPolicy.objects.filter(is_active=True))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Сотрудники"

    bold = Font(bold=True)
    for column, name in enumerate(EMPLOYEE_CSV_COLUMNS, start=1):
        sheet.cell(row=1, column=column, value=name).font = bold

    for row_index, row in enumerate(rows, start=2):
        masked = mask_employee_data(
            row, user_permissions=user_permissions, policies=policies
        )
        for column, name in enumerate(EMPLOYEE_CSV_COLUMNS, start=1):
            value = sanitize_cell(normalize_value(masked.get(name)))
            sheet.cell(row=row_index, column=column, value=value)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
