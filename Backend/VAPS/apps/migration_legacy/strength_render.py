"""Pure rendering of a strength report (1.7) — text table + simple .xlsx.

No ORM, no Django: tested without a database, eats the
``StrengthReportResult`` of story 1.7 as is. The renderer NEVER prints
(the command prints) and NEVER reads the wall clock — every value comes
from the result. ``openpyxl`` is imported INSIDE ``build_workbook`` so the
text-table path (and this module's import) work without openpyxl
installed.

"Simple" is deliberate (Решение №3): one sheet per date, a header, the
division rows and a totals row — no styles, formulas, merges or
golden-master. The full document generator lives in E6 (6.3/6.8).
"""

from apps.operations.statuses.services import REPORT_COLUMNS

# Display strings for stdout/sheet headers. These are NOT domain terms —
# the canon stays the column CODE; the Glossary is not extended (story note
# on Task 2). Order of the per-status columns follows REPORT_COLUMNS.
COLUMN_LABELS_RU = {
    "staff_total": "Штат",
    "list_total": "Список",
    "vacancies": "Вакансии",
    "SICK": "Больничный",
    "VACATION": "Отпуск",
    "COMMAND": "Командировка",
    "TRAINING": "Учёба",
    "OTHER": "Прочие",
    "DETACHED": "Откомандировано",
    "AFTER_DUTY": "После деж.",
    "BEFORE_DUTY": "Перед деж.",
    "ON_DUTY": "На деж.",
    "IN_SERVICE": "В строю",
    "ATTACHED": "Прикомандировано",
}

# Fixed left-to-right layout of every rendered row, by logical key.
_LAYOUT = (
    "name",
    "staff_total",
    "list_total",
    "vacancies",
    *REPORT_COLUMNS,
    "ATTACHED",
)

# Header label for the leftmost (division name) column.
_NAME_HEADER = "Подразделение"
_TOTALS_LABEL = "Итого"


def _header_cells():
    cells = [_NAME_HEADER]
    for key in _LAYOUT[1:]:
        cells.append(COLUMN_LABELS_RU[key])
    return cells


def _row_cells(name, staff_total, list_total, vacancies, columns, attached):
    """One logical row as a list of cell values (ints + the name string)."""
    cells = [name, staff_total, list_total, vacancies]
    cells.extend(columns[column] for column in REPORT_COLUMNS)
    cells.append(attached)
    return cells


def render_table(result):
    """Text table for stdout: a row per division + a totals row.

    Numbers are right-aligned; ``attached`` shows as ``+N``. Row order is
    the order of ``result.rows`` (already sorted in ``derive_report`` — the
    renderer must not re-sort). ``violations`` and ``warnings`` are printed
    below the table, human-readably. Returns a string; prints nothing.
    """
    header = _header_cells()
    body = []
    for row in result.rows:
        body.append(
            _row_cells(
                row.name,
                row.staff_total,
                row.list_total,
                row.vacancies,
                row.columns,
                row.attached,
            )
        )
    totals = result.totals
    totals_row = _row_cells(
        _TOTALS_LABEL,
        totals.staff_total,
        totals.list_total,
        totals.vacancies,
        totals.columns,
        totals.attached,
    )

    # Column widths from the widest rendered cell (the attached column shows
    # "+N", everything else is str()).
    def display(value, key):
        if key == "ATTACHED":
            return f"+{value}"
        return str(value)

    all_rows = [header, *body, totals_row]
    widths = [0] * len(_LAYOUT)
    for cells in all_rows:
        for index, (value, key) in enumerate(zip(cells, _LAYOUT)):
            text = value if index == 0 else display(value, key)
            widths[index] = max(widths[index], len(str(text)))

    def fmt(cells, is_header):
        parts = []
        for index, (value, key) in enumerate(zip(cells, _LAYOUT)):
            if index == 0:
                parts.append(str(value).ljust(widths[index]))
            elif is_header:
                parts.append(str(value).rjust(widths[index]))
            else:
                parts.append(display(value, key).rjust(widths[index]))
        return "  ".join(parts).rstrip()

    lines = [f"Расход на {result.business_date.isoformat()}", ""]
    lines.append(fmt(header, is_header=True))
    for cells in body:
        lines.append(fmt(cells, is_header=False))
    lines.append(fmt(totals_row, is_header=False))

    if result.violations:
        lines.append("")
        lines.append("Нарушения:")
        for violation in result.violations:
            lines.append(
                f"  - {violation['reason']}: подразделение "
                f"{violation['division_id']} "
                f"(Штат {violation['staff_total']} < Список "
                f"{violation['list_total']})"
            )
    if result.warnings:
        lines.append("")
        lines.append("Предупреждения:")
        for warning in result.warnings:
            lines.append(
                f"  ~ {warning['reason']}: подразделение "
                f"{warning['division_id']}"
            )
    return "\n".join(lines)


def build_workbook(results):
    """``list[StrengthReportResult]`` -> ``openpyxl.Workbook`` (one sheet/date).

    Sheet name = ISO date (10 chars, free of openpyxl's forbidden title
    characters ``/\\?*[]:``; unique dates => no collision). The default
    ``Sheet`` openpyxl seeds on ``Workbook()`` is removed so the file holds
    exactly one sheet per date (the test counts sheets). Imported here, not
    at module level, so the text-table path needs no openpyxl.
    """
    from openpyxl import Workbook

    workbook = Workbook()
    # Drop the auto-created default sheet; we add our own named ones.
    default = workbook.active
    workbook.remove(default)

    header = _header_cells()
    for result in results:
        sheet = workbook.create_sheet(title=result.business_date.isoformat())
        sheet.append(header)
        for row in result.rows:
            sheet.append(
                _row_cells(
                    row.name,
                    row.staff_total,
                    row.list_total,
                    row.vacancies,
                    row.columns,
                    row.attached,
                )
            )
        totals = result.totals
        sheet.append(
            _row_cells(
                _TOTALS_LABEL,
                totals.staff_total,
                totals.list_total,
                totals.vacancies,
                totals.columns,
                totals.attached,
            )
        )
    return workbook
