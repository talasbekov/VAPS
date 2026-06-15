"""Unit tests for the pure strength-report renderer (no DB, gate).

Builds StrengthReportResult by hand (1.7 contract) and asserts the text
table and the simple .xlsx workbook — both pure functions of the result.
"""

import io
from datetime import date

from openpyxl import load_workbook

from apps.migration_legacy.strength_render import (
    COLUMN_LABELS_RU,
    build_workbook,
    render_table,
)
from apps.operations.statuses.services import (
    REPORT_COLUMNS,
    DivisionReportRow,
    ReportTotals,
    StrengthReportResult,
)


def make_row(name, staff, list_total, vacancies, columns=None, attached=0):
    cols = {column: 0 for column in REPORT_COLUMNS}
    cols.update(columns or {})
    return DivisionReportRow(
        division_id=name,
        name=name,
        staff_total=staff,
        list_total=list_total,
        vacancies=vacancies,
        columns=cols,
        attached=attached,
    )


def make_result(on_date, rows, violations=None, warnings=None):
    totals_columns = {column: 0 for column in REPORT_COLUMNS}
    staff = lst = vac = att = 0
    for row in rows:
        for column in REPORT_COLUMNS:
            totals_columns[column] += row.columns[column]
        staff += row.staff_total
        lst += row.list_total
        vac += row.vacancies
        att += row.attached
    totals = ReportTotals(
        staff_total=staff,
        list_total=lst,
        vacancies=vac,
        columns=totals_columns,
        attached=att,
    )
    return StrengthReportResult(
        business_date=on_date,
        rows=rows,
        totals=totals,
        violations=violations or [],
        warnings=warnings or [],
    )


class TestRenderTable:
    def test_table_has_division_rows_and_totals(self):
        rows = [
            make_row("DEP1", 5, 1, 4, {"VACATION": 1}),
            make_row("DIR1", 1, 1, 0, {"TRAINING": 1}),
        ]
        text = render_table(make_result(date(2026, 6, 4), rows))
        assert "2026-06-04" in text
        assert "DEP1" in text and "DIR1" in text
        # RU labels in the header, not raw column codes.
        assert COLUMN_LABELS_RU["VACATION"] in text
        assert COLUMN_LABELS_RU["IN_SERVICE"] in text
        # A totals line is present.
        assert "Итого" in text

    def test_attached_rendered_as_plus_n(self):
        rows = [make_row("DEP1", 3, 2, 1, {"IN_SERVICE": 2}, attached=2)]
        text = render_table(make_result(date(2026, 6, 4), rows))
        assert "+2" in text

    def test_violations_and_warnings_listed(self):
        rows = [make_row("DEP1", 0, 1, 0, {"IN_SERVICE": 1})]
        result = make_result(
            date(2026, 6, 4),
            rows,
            violations=[
                {
                    "division_id": "DEP1",
                    "reason": "staff_lt_list",
                    "staff_total": 0,
                    "list_total": 1,
                }
            ],
            warnings=[{"division_id": "DEP1", "reason": "no_staffing_record"}],
        )
        text = render_table(result)
        assert "staff_lt_list" in text
        assert "no_staffing_record" in text

    def test_row_order_is_preserved(self):
        rows = [
            make_row("B-div", 1, 0, 1),
            make_row("A-div", 1, 0, 1),
        ]
        text = render_table(make_result(date(2026, 6, 4), rows))
        # derive_report already sorted result.rows; the renderer must NOT
        # re-sort — B before A here.
        assert text.index("B-div") < text.index("A-div")

    def test_render_table_does_not_print(self, capsys):
        rows = [make_row("DEP1", 1, 1, 0, {"IN_SERVICE": 1})]
        render_table(make_result(date(2026, 6, 4), rows))
        captured = capsys.readouterr()
        assert captured.out == ""


class TestBuildWorkbook:
    def test_one_sheet_per_date_no_default_sheet(self):
        results = [
            make_result(d, [make_row("DEP1", 5, 1, 4, cols)])
            for d, cols in (
                (date(2026, 6, 4), {"VACATION": 1}),
                (date(2026, 6, 5), {"VACATION": 1}),
                (date(2026, 6, 6), {"IN_SERVICE": 1}),
            )
        ]
        wb = build_workbook(results)
        # Exactly one sheet per date — no leftover default "Sheet".
        assert wb.sheetnames == ["2026-06-04", "2026-06-05", "2026-06-06"]

    def test_sheet_has_header_and_numbers(self):
        result = make_result(
            date(2026, 6, 4),
            [make_row("DEP1", 5, 1, 4, {"VACATION": 1})],
        )
        wb = build_workbook([result])
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        reloaded = load_workbook(buffer)
        sheet = reloaded["2026-06-04"]
        header = [cell.value for cell in sheet[1]]
        assert COLUMN_LABELS_RU["staff_total"] in header
        assert COLUMN_LABELS_RU["VACATION"] in header
        # The data row carries the division name and its Штат.
        body = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]
        flat = [value for row in body for value in row]
        assert "DEP1" in flat
        assert 5 in flat  # staff_total

    def test_totals_row_present(self):
        result = make_result(
            date(2026, 6, 4),
            [
                make_row("DEP1", 5, 1, 4, {"VACATION": 1}),
                make_row("DIR1", 1, 1, 0, {"TRAINING": 1}),
            ],
        )
        wb = build_workbook([result])
        sheet = wb["2026-06-04"]
        col_a = [row[0].value for row in sheet.iter_rows()]
        assert "Итого" in col_a
