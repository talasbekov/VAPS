"""Integration test for the strength_report command (Postgres, gate).

Imports the real donor slice, then exercises the command end to end:
text tables, the simple .xlsx (one sheet per inclusive day), the
categorized diff against the synthetic baseline, and the mechanical DoD
gate — unclassified / data-loss discrepancies exit non-zero (AC-5).
"""

import io
import json
from pathlib import Path

import pytest
from django.core.management import call_command
from django.core.management.base import CommandError
from openpyxl import load_workbook

FIXTURES = Path(__file__).parent / "fixtures"
SLICE = FIXTURES / "donor_slice.json"
SAMPLE = FIXTURES / "donor_baseline_sample.json"

pytestmark = pytest.mark.django_db


def import_slice():
    call_command(
        "import_donor_slice",
        str(SLICE),
        "--days",
        "7",
        "--until",
        "2026-06-07",
        stdout=io.StringIO(),
    )


def report(*extra, stdout=None):
    out = stdout or io.StringIO()
    call_command("strength_report", *extra, stdout=out)
    return out.getvalue()


class TestTextAndXlsx:
    def test_range_prints_tables_per_division(self):
        import_slice()
        text = report("--from", "2026-06-04", "--to", "2026-06-06")
        assert "2026-06-04" in text
        assert "2026-06-05" in text
        assert "2026-06-06" in text
        # The text table renders Division NAMES (row.name); codes appear in
        # the diff report, not here.
        assert "Департамент кадров" in text and "Управление учёта" in text

    def test_xlsx_has_exactly_one_sheet_per_inclusive_day(self, tmp_path):
        import_slice()
        path = tmp_path / "rashod.xlsx"
        report("--from", "2026-06-04", "--to", "2026-06-06", "--xlsx", str(path))
        wb = load_workbook(path)
        # --from D --to D+2 inclusive on both ends => 3 sheets, ISO names.
        assert wb.sheetnames == ["2026-06-04", "2026-06-05", "2026-06-06"]

    def test_single_day_is_one_sheet(self, tmp_path):
        import_slice()
        path = tmp_path / "one.xlsx"
        report("--date", "2026-06-04", "--xlsx", str(path))
        wb = load_workbook(path)
        assert wb.sheetnames == ["2026-06-04"]


class TestDiffAndGate:
    def test_sample_diff_prints_categories_and_blocks_gate(self):
        import_slice()
        out = io.StringIO()
        # The sample carries gate-blocking discrepancies — a data-loss surplus
        # (DEP1 IN_SERVICE on 06-05) and an unexplained 1:1-column donor
        # surplus (DIR1 DETACHED on 06-04) — so the gate must fire AFTER
        # printing the categorized diff (AC-2 + AC-5). NB: code review
        # 2026-06-16 (finding C1) removed the unsound auto model/single_winner
        # label; an ambiguous donor type-column surplus now stays gate-blocking
        # rather than being green-lit.
        with pytest.raises(CommandError):
            report(
                "--from",
                "2026-06-04",
                "--to",
                "2026-06-06",
                "--diff-baseline",
                str(SAMPLE),
                stdout=out,
            )
        text = out.getvalue()
        assert "model/aggregator_inferred" in text
        assert "model/attached_source" in text
        assert "timing/half_open_end" in text
        assert "data/skipped_employee" in text
        assert "[unclassified]" in text
        assert "model/single_winner" not in text  # unsound label, removed (C1)
        assert "UNCLASSIFIED" in text

    def test_planted_unclassified_exits_nonzero(self, tmp_path):
        import_slice()
        # A one-day baseline with an unexplained VAPS surplus (donor has
        # fewer in a 1:1 column) — no structural rule explains it.
        baseline = {
            "days": [
                {
                    "date": "2026-06-04",
                    "rows": [
                        {
                            "division_code": "DEP1",
                            "division_name": "Департамент кадров",
                            "staff_unit": 5,
                            "in_service": 0,
                            "vacation": 0,
                            "sick_leave": 0,
                            "business_trip": 0,
                            "training": 0,
                            "seconded_in": 0,
                            "seconded_out": 0,
                            "other_absence": 0,
                        }
                    ],
                }
            ]
        }
        path = tmp_path / "planted.json"
        path.write_text(json.dumps(baseline), encoding="utf-8")
        out = io.StringIO()
        with pytest.raises(CommandError):
            report("--date", "2026-06-04", "--diff-baseline", str(path), stdout=out)
        assert "UNCLASSIFIED" in out.getvalue()

    def test_missing_day_in_baseline_is_command_error(self, tmp_path):
        import_slice()
        baseline = {"days": [{"date": "2026-06-04", "rows": []}]}
        path = tmp_path / "short.json"
        path.write_text(json.dumps(baseline), encoding="utf-8")
        with pytest.raises(CommandError, match="no day 2026-06-05"):
            report(
                "--from",
                "2026-06-04",
                "--to",
                "2026-06-05",
                "--diff-baseline",
                str(path),
                stdout=io.StringIO(),
            )


class TestArgumentValidation:
    def test_nonexistent_division_errors_before_compute(self):
        import_slice()
        with pytest.raises(CommandError, match="not found"):
            report(
                "--date",
                "2026-06-04",
                "--division",
                "00000000-0000-0000-0000-000000000000",
                stdout=io.StringIO(),
            )

    def test_malformed_division_is_command_error_not_traceback(self):
        # Finding C5: a non-UUID --division raises ValidationError at
        # filter-build time; the CLI boundary must convert it to CommandError,
        # not leak a Django traceback.
        with pytest.raises(CommandError, match="not a valid UUID"):
            report(
                "--date",
                "2026-06-04",
                "--division",
                "not-a-uuid",
                stdout=io.StringIO(),
            )

    def test_date_and_from_are_mutually_exclusive(self):
        with pytest.raises(CommandError, match="mutually exclusive"):
            report(
                "--date",
                "2026-06-04",
                "--from",
                "2026-06-04",
                stdout=io.StringIO(),
            )

    def test_from_without_to_errors(self):
        with pytest.raises(CommandError):
            report("--from", "2026-06-04", stdout=io.StringIO())

    def test_neither_date_nor_range_errors(self):
        with pytest.raises(CommandError):
            report(stdout=io.StringIO())

    def test_from_after_to_errors(self):
        with pytest.raises(CommandError, match="must not be after"):
            report("--from", "2026-06-06", "--to", "2026-06-04", stdout=io.StringIO())

    def test_invalid_date_errors(self):
        with pytest.raises(CommandError, match="not a date"):
            report("--date", "31.05.2026", stdout=io.StringIO())
