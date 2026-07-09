"""Story 6.4 — чистый рендерер официального .xlsx расхода.

Канон вёрстки — addendum §8, тот же, что у .docx 6.3 (XLSX — документная
форма, Д3): лист = ISO-дата бизнес-дня (свободна от запрещённых openpyxl
символов ``/\\?*[]:`` — прецедент ``strength_render``), титул merge'ится на
ширину таблицы (16pt, жирный, Times New Roman), шапка — литеральный дубль
17-колоночной шапки docx (12pt, жирный), печать A4 landscape.

Ячейки тела: безчленные числовые — **int** (числа как числа, AC-2); статусные
с членами — ТЕКСТ ``{count}\\n{строки членов}`` (строки — ``_member_line`` из
expense_docx, cap ``CELL_MAX_MEMBERS`` + хвост ``… ещё N``, wrap_text);
ATTACHED — ``+N`` (+ члены) — прикомандированные вне «По списку». Один шрифт
на ячейку — ограничение формата (rich text не используется), потому члены
набраны тем же 12pt, а не 8pt docx. Строка ИТОГО — жирным, зеркало docx:
«№» пуст, лейбл — в колонке «Управление», далее числа, ATTACHED — ``+N``.

Рендерер — ЧИСТАЯ функция ``ExpenseDocumentData -> bytes``: без ORM, без
wall-clock, без сети и записи на диск; ``openpyxl`` импортируется ЛЕНИВО
внутри ``generate_expense_xlsx`` (зеркало 6.3). Рендерер НЕ считает, НЕ
пересортировывает строки и формулы сходимости НЕ проверяет (зона derive и
рантайм-ассерта выпуска 6.5); Excel-формул в файле нет — только
предвычисленные значения.
"""

import io

from apps.documents.generators.expense_docx import (
    _DATE_FORMAT,
    _FIXED_HEAD,
    _TOTALS_LABEL,
    _member_line,
    CELL_MAX_MEMBERS,
    DOCX_COLUMN_LABELS,
    DOCX_COLUMNS,
    FONT_NAME,
    TABLE_SIZE_PT,
    TITLE_SIZE_PT,
)

_TABLE_WIDTH = len(_FIXED_HEAD) + len(DOCX_COLUMNS)


def _status_cell_value(key, cell_data):
    """Значение статусной ячейки: int без членов, текст «count + члены» с
    членами; ATTACHED — всегда текст ``+N``. Усечение cap + ``… ещё N`` —
    дубль closure ``fill_status_cell`` (импортировать её нельзя, Ловушка №3)."""
    count_text = f"+{cell_data.count}" if key == "ATTACHED" else str(cell_data.count)
    if not cell_data.members:
        return count_text if key == "ATTACHED" else cell_data.count
    shown = cell_data.members[:CELL_MAX_MEMBERS]
    lines = [_member_line(member) for member in shown]
    hidden = len(cell_data.members) - len(shown)
    if hidden > 0:
        lines.append(f"… ещё {hidden}")
    return "\n".join([count_text, *lines])


def generate_expense_xlsx(data):
    """``ExpenseDocumentData`` → байты официального .xlsx (канон §8).

    Ленивая загрузка openpyxl — зеркало ``generate_expense_docx``.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    plain = Font(name=FONT_NAME, size=TABLE_SIZE_PT)
    bold = Font(name=FONT_NAME, size=TABLE_SIZE_PT, bold=True)
    member_alignment = Alignment(wrap_text=True, vertical="top")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = data.business_date.isoformat()

    title = sheet.cell(row=1, column=1)
    title.value = (
        f"{data.division_title} ЖЕКЕ ҚҰРАМЫНЫҢ САПТЫҚ ТІЗІМІ "
        f"{data.business_date.strftime(_DATE_FORMAT)} ЖЫЛҒЫ"
    )
    title.font = Font(name=FONT_NAME, size=TITLE_SIZE_PT, bold=True)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_TABLE_WIDTH)

    header = _FIXED_HEAD + tuple(DOCX_COLUMN_LABELS[key] for key in DOCX_COLUMNS)
    for column, label in enumerate(header, start=1):
        cell = sheet.cell(row=2, column=column, value=label)
        cell.font = bold

    for number, row in enumerate(data.rows, start=1):
        sheet_row = 2 + number
        values = [number, row.name, row.staff_total, row.list_total, row.vacancies]
        for column, value in enumerate(values, start=1):
            sheet.cell(row=sheet_row, column=column, value=value).font = plain
        for offset, key in enumerate(DOCX_COLUMNS, start=len(_FIXED_HEAD) + 1):
            cell_data = row.attached if key == "ATTACHED" else row.cells[key]
            cell = sheet.cell(
                row=sheet_row, column=offset, value=_status_cell_value(key, cell_data)
            )
            cell.font = plain
            if cell_data.members:
                cell.alignment = member_alignment

    totals = data.totals
    totals_sheet_row = 2 + len(data.rows) + 1
    totals_values = [
        _TOTALS_LABEL,
        totals.staff_total,
        totals.list_total,
        totals.vacancies,
    ]
    for column, value in enumerate(totals_values, start=2):
        sheet.cell(row=totals_sheet_row, column=column, value=value).font = bold
    for offset, key in enumerate(DOCX_COLUMNS, start=len(_FIXED_HEAD) + 1):
        if key == "ATTACHED":
            value = f"+{totals.attached}"
        else:
            value = totals.columns[key]
        sheet.cell(row=totals_sheet_row, column=offset, value=value).font = bold

    sheet.page_setup.orientation = "landscape"
    # Константа живёт на Worksheet, НЕ на PageSetup; строковый сет
    # ``paperSize = "A4"`` — TypeError (инкантация AC-2, проверено в venv).
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
