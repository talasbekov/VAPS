"""Тесты чистого рендерера .xlsx расхода (Story 6.4, AC-1/2/6).

БЕЗ ``django_db``: рендерер — чистая функция (datacls-вход → bytes), читаем
сгенерированное обратно через ``load_workbook(BytesIO(result))``. Фикстуры —
копия паттерна ``test_expense_docx_generator.py`` (тот же контракт входа).
Числа входа — литералы: рендерер формулы НЕ проверяет (зона derive/6.5).
"""

from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook

from apps.documents.generators import (
    CELL_MAX_MEMBERS,
    DOCX_COLUMN_LABELS,
    DOCX_COLUMNS,
    FONT_NAME,
    ExpenseCell,
    ExpenseCellMember,
    ExpenseDocumentData,
    ExpenseRow,
    ExpenseTotals,
    generate_expense_xlsx,
)

D = date(2026, 7, 8)

# Раскладка листа: строка 1 — титул (merge на ширину таблицы), строка 2 —
# шапка, строки 3.. — данные, последняя — ИТОГО. Колонки 1-based:
# № | Управление | По штату | По списку | Вакансии | 12 колонок DOCX_COLUMNS.
_FIXED_HEAD = ("№", "Управление", "По штату", "По списку", "Вакансии")
_TITLE_ROW = 1
_HEADER_ROW = 2
_FIRST_DATA_ROW = 3


def _col(key):
    return 1 + len(_FIXED_HEAD) + DOCX_COLUMNS.index(key)


def _member(rank="капитан", full_name="Иванов Иван", start=None, end=None):
    return ExpenseCellMember(
        rank=rank,
        full_name=full_name,
        date_start=start or date(2026, 7, 1),
        date_end=end or date(2026, 7, 10),
    )


def _cells(**counts):
    """Все 11 статусных ячеек (без ATTACHED); именованные — с содержимым."""
    cells = {}
    for key in DOCX_COLUMNS:
        if key == "ATTACHED":
            continue
        cells[key] = counts.get(key, ExpenseCell(count=0, members=()))
    return cells


def _totals(**overrides):
    columns = {key: 0 for key in DOCX_COLUMNS if key != "ATTACHED"}
    columns.update(overrides.pop("columns", {}))
    defaults = {"staff_total": 10, "list_total": 8, "vacancies": 2, "attached": 1}
    defaults.update(overrides)
    return ExpenseTotals(columns=columns, **defaults)


def _data(rows=None, totals=None, division_title="1-е управление"):
    if rows is None:
        rows = [
            ExpenseRow(
                name="1-е управление",
                staff_total=10,
                list_total=8,
                vacancies=2,
                cells=_cells(
                    IN_SERVICE=ExpenseCell(count=4, members=()),
                    ON_DUTY=ExpenseCell(count=1, members=(_member(),)),
                    TRAINING=ExpenseCell(
                        count=2,
                        members=(
                            _member(full_name="Петров Пётр"),
                            _member(rank="майор", full_name="Сидоров Сидор"),
                        ),
                    ),
                    VACATION=ExpenseCell(count=1, members=(_member(rank=""),)),
                ),
                attached=ExpenseCell(count=1, members=(_member(),)),
            )
        ]
    if totals is None:
        totals = _totals(
            columns={"IN_SERVICE": 4, "ON_DUTY": 1, "TRAINING": 2, "VACATION": 1}
        )
    return ExpenseDocumentData(
        division_title=division_title, business_date=D, rows=rows, totals=totals
    )


def _generated(data):
    result = generate_expense_xlsx(data)
    assert isinstance(result, bytes)
    return load_workbook(BytesIO(result))


def test_bytes_roundtrip_single_sheet_named_iso_date():
    result = generate_expense_xlsx(_data())
    assert isinstance(result, bytes)
    assert result[:2] == b"PK"
    workbook = load_workbook(BytesIO(result))
    assert workbook.sheetnames == ["2026-07-08"]  # ровно один лист, ISO-дата


def test_title_text_style_and_merge_across_table_width():
    sheet = _generated(_data()).active
    title = sheet.cell(row=_TITLE_ROW, column=1)
    assert title.value == (
        "1-е управление ЖЕКЕ ҚҰРАМЫНЫҢ САПТЫҚ ТІЗІМІ 08.07.2026 ЖЫЛҒЫ"
    )
    assert title.font.name == FONT_NAME
    assert title.font.size == 16
    assert title.font.bold is True
    merges = [str(merged) for merged in sheet.merged_cells.ranges]
    assert "A1:Q1" in merges  # merge по ширине 17-колоночной таблицы


def test_header_row_full_label_order():
    sheet = _generated(_data()).active
    labels = [sheet.cell(row=_HEADER_ROW, column=c).value for c in range(1, 18)]
    assert labels == list(_FIXED_HEAD) + [
        DOCX_COLUMN_LABELS[key] for key in DOCX_COLUMNS
    ]
    # Литеральная сверка канона AC-2 — от дрейфа констант (зеркало docx-теста).
    assert labels == [
        "№",
        "Управление",
        "По штату",
        "По списку",
        "Вакансии",
        "В строю",
        "На дежурстве",
        "После дежурства",
        "В командировке",
        "Учёба/соревнования/конференция",
        "В отпуске",
        "На больничном",
        "Прикомандирован",
        "Откомандирован",
        "Перед дежурством",
        "Иное",
        "Уточняется",
    ]
    for column in range(1, 18):
        font = sheet.cell(row=_HEADER_ROW, column=column).font
        assert font.name == FONT_NAME
        assert font.size == 12
        assert font.bold is True


def test_body_row_values_and_memberless_cells_are_int():
    sheet = _generated(_data()).active
    row = _FIRST_DATA_ROW
    assert sheet.cell(row=row, column=1).value == 1  # № — int с 1
    assert sheet.cell(row=row, column=2).value == "1-е управление"
    for column, expected in ((3, 10), (4, 8), (5, 2)):
        value = sheet.cell(row=row, column=column).value
        assert value == expected
        assert isinstance(value, int)  # числа как числа (AC-2)
    in_service = sheet.cell(row=row, column=_col("IN_SERVICE")).value
    assert in_service == 4
    assert isinstance(in_service, int)
    sick = sheet.cell(row=row, column=_col("SICK")).value
    assert sick == 0
    assert isinstance(sick, int)


def test_member_cell_is_text_count_first_line_with_wrap():
    sheet = _generated(_data()).active
    cell = sheet.cell(row=_FIRST_DATA_ROW, column=_col("TRAINING"))
    assert cell.value == (
        "2\n"
        "капитан Петров Пётр — 01.07.2026–10.07.2026\n"
        "майор Сидоров Сидор — 01.07.2026–10.07.2026"
    )
    assert cell.value.splitlines()[0] == "2"
    assert cell.alignment.wrap_text is True
    assert cell.alignment.vertical == "top"


def test_member_line_empty_rank_has_no_leading_space():
    sheet = _generated(_data()).active
    cell = sheet.cell(row=_FIRST_DATA_ROW, column=_col("VACATION"))
    assert cell.value.splitlines()[1] == "Иванов Иван — 01.07.2026–10.07.2026"


def test_attached_rendered_as_plus_n_with_members():
    sheet = _generated(_data()).active
    cell = sheet.cell(row=_FIRST_DATA_ROW, column=_col("ATTACHED"))
    lines = cell.value.splitlines()
    assert lines[0] == "+1"
    assert lines[1] == "капитан Иванов Иван — 01.07.2026–10.07.2026"


def test_truncation_over_cap_members():
    members = tuple(
        _member(full_name=f"Сотрудник {n:02d}") for n in range(CELL_MAX_MEMBERS + 1)
    )
    row = ExpenseRow(
        name="Отдел",
        staff_total=30,
        list_total=25,
        vacancies=5,
        cells=_cells(ON_DUTY=ExpenseCell(count=len(members), members=members)),
        attached=ExpenseCell(count=0, members=()),
    )
    sheet = _generated(_data(rows=[row], totals=_totals())).active
    lines = sheet.cell(row=_FIRST_DATA_ROW, column=_col("ON_DUTY")).value.splitlines()
    assert lines[0] == str(CELL_MAX_MEMBERS + 1)
    assert len(lines) == 1 + CELL_MAX_MEMBERS + 1  # count + 20 членов + хвост
    assert "Сотрудник 00" in lines[1]
    assert lines[-1] == "… ещё 1"


def test_totals_row_values_and_bold():
    sheet = _generated(_data()).active
    totals_row = sheet.max_row
    assert sheet.cell(row=totals_row, column=1).value is None  # № пуст
    assert sheet.cell(row=totals_row, column=2).value == "ИТОГО"
    assert sheet.cell(row=totals_row, column=3).value == 10
    assert sheet.cell(row=totals_row, column=4).value == 8
    assert sheet.cell(row=totals_row, column=5).value == 2
    in_service = sheet.cell(row=totals_row, column=_col("IN_SERVICE")).value
    assert in_service == 4
    assert isinstance(in_service, int)
    assert sheet.cell(row=totals_row, column=_col("ATTACHED")).value == "+1"
    for column in range(2, 18):
        font = sheet.cell(row=totals_row, column=column).font
        assert font.name == FONT_NAME
        assert font.size == 12
        assert font.bold is True


def test_page_setup_landscape_a4():
    sheet = _generated(_data()).active
    assert sheet.page_setup.orientation == "landscape"
    # PAPERSIZE_A4 — константа Worksheet ('9'); после round-trip читается int 9.
    assert int(sheet.page_setup.paperSize) == 9


def test_rows_rendered_in_given_order_without_resort():
    rows = [
        ExpenseRow(
            name=name,
            staff_total=1,
            list_total=1,
            vacancies=0,
            cells=_cells(IN_SERVICE=ExpenseCell(count=1, members=())),
            attached=ExpenseCell(count=0, members=()),
        )
        for name in ("Яблоко", "Абрикос")  # нарочно НЕ по алфавиту
    ]
    sheet = _generated(_data(rows=rows, totals=_totals())).active
    assert sheet.cell(row=_FIRST_DATA_ROW, column=1).value == 1
    assert sheet.cell(row=_FIRST_DATA_ROW, column=2).value == "Яблоко"
    assert sheet.cell(row=_FIRST_DATA_ROW + 1, column=1).value == 2
    assert sheet.cell(row=_FIRST_DATA_ROW + 1, column=2).value == "Абрикос"


def test_attached_without_members_stays_text_plus_n():
    # ATTACHED — всегда текст «+N», даже без членов (в отличие от прочих
    # безчленных ячеек, которые обязаны быть int): паритет с docx/PDF.
    row = ExpenseRow(
        name="Отдел",
        staff_total=5,
        list_total=4,
        vacancies=1,
        cells=_cells(),
        attached=ExpenseCell(count=3, members=()),
    )
    sheet = _generated(_data(rows=[row], totals=_totals())).active
    value = sheet.cell(row=_FIRST_DATA_ROW, column=_col("ATTACHED")).value
    assert value == "+3"
    assert isinstance(value, str)


def test_zero_rows_renders_header_and_totals_only():
    # Row-агностичность (Д11, зеркало QA-судьи 6.3 по docx): 0 строк →
    # титул + шапка + ИТОГО без падения; опора контракта для свода 6.5.
    sheet = _generated(_data(rows=[], totals=_totals())).active
    assert sheet.max_row == 3  # титул + шапка + ИТОГО
    assert sheet.cell(row=3, column=1).value is None
    assert sheet.cell(row=3, column=2).value == "ИТОГО"


def test_missing_status_key_raises_loudly_stop_semantics():
    # STOP-семантика (AC-1): неполный контракт (ячейка колонки пропала) роняет
    # рендерер KeyError'ом, а не рендерит документ с молчаливой дырой.
    cells = _cells()
    del cells["SICK"]
    row = ExpenseRow(
        name="Отдел",
        staff_total=1,
        list_total=1,
        vacancies=0,
        cells=cells,
        attached=ExpenseCell(count=0, members=()),
    )
    with pytest.raises(KeyError):
        generate_expense_xlsx(_data(rows=[row], totals=_totals()))
