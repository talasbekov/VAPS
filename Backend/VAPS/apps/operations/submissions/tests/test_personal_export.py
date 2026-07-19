"""Story 10.8 — unit-тесты чистого билдера личного экспорта («щит»).

Задача сюиты — доказать, что файл ОТРАЖАЕТ снапшот, а не пересчитывает его.
Билдер чистый (ни ORM, ни часов), поэтому `django_db` здесь не нужен: книга
читается обратно через ``openpyxl.load_workbook`` и проверяется по ячейкам —
байтовый факт, а не мок.

Ключевой гвард сюиты — `test_export_modules_carry_no_derived_default`: он
читает ИСХОДНИК прод-модулей и требует отсутствия литералов производного
состояния. Гвард именно на исходник, а не на вывод: в выводе такой статус
появляться ВПРАВЕ, когда он пришёл строкой снапшота (AC-4).
"""

import io
from datetime import date, datetime, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

import pytest
from django.conf import settings
from openpyxl import load_workbook

from apps.operations.submissions import personal_export
from apps.operations.submissions.personal_export import build_personal_export_xlsx
from apps.operations.submissions.services import personal_export_service

TODAY = date(2026, 6, 4)

EMP_A = "11111111-1111-1111-1111-111111111111"
EMP_B = "22222222-2222-2222-2222-222222222222"
EMP_C = "33333333-3333-3333-3333-333333333333"
ORPHAN = "99999999-9999-9999-9999-999999999999"

TABLE_COLUMNS = (
    "№",
    "Звание",
    "ФИО",
    "Статус",
    "Код статуса",
    "С",
    "По",
    "Источник",
)

STATUS_NAMES = {
    "SICK_LEAVE": "Больничный",
    "DUTY": "Наряд",
    "IN_SERVICE": "В строю",
}


def _member(employee_id, full_name, rank):
    return {"employee_id": employee_id, "full_name": full_name, "rank": rank}


def _row(employee_id, code, *, status_id=1, start="2026-06-01", end="2026-06-10"):
    return {
        "employee_id": employee_id,
        "status_type_code": code,
        "status_id": status_id,
        "date_start": start,
        "date_end": end,
        "source": "USER",
    }


def _snapshot(*, roster=None, rows=None, schema_version=1):
    if roster is None:
        roster = [
            _member(EMP_A, "Абдиров А.А.", "капитан"),
            _member(EMP_B, "Бекова Б.Б.", "лейтенант"),
        ]
    return {
        "schema_version": schema_version,
        "roster": roster,
        "rows": [] if rows is None else rows,
    }


def _build(**overrides):
    kwargs = {
        "snapshot": _snapshot(),
        "division_title": "Управление А",
        "business_date": TODAY,
        "version": 3,
        "is_current": True,
        "event_label": "С изменениями",
        "submitted_by": "seed-op",
        "submitted_at_label": "04.06.2026 17:05",
        "late": False,
        "status_names": STATUS_NAMES,
    }
    kwargs.update(overrides)
    return build_personal_export_xlsx(**kwargs)


def _sheet(payload):
    return load_workbook(io.BytesIO(payload)).active


def _header_row_index(sheet):
    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if tuple(row[: len(TABLE_COLUMNS)]) == TABLE_COLUMNS:
            return index
    raise AssertionError("шапка таблицы состава не найдена")


def _passport(sheet):
    """{лейбл: значение} блока-шапки — колонки A/B до шапки таблицы."""
    stop = _header_row_index(sheet)
    out = {}
    for row in sheet.iter_rows(
        min_row=1, max_row=stop - 1, max_col=2, values_only=True
    ):
        if row[0] is not None:
            out[row[0]] = row[1]
    return out


def _table(sheet):
    start = _header_row_index(sheet) + 1
    rows = []
    for row in sheet.iter_rows(
        min_row=start, max_col=len(TABLE_COLUMNS), values_only=True
    ):
        # Пустая строка-разделитель отделяет таблицу от подвала (итог/легенда).
        if row[0] is None:
            break
        rows.append(row)
    return rows


def _footer_text(sheet):
    start = _header_row_index(sheet) + len(_table(sheet)) + 1
    return [
        row[0]
        for row in sheet.iter_rows(min_row=start, max_col=1, values_only=True)
        if row[0] is not None
    ]


# --- AC-2: паспорт сдачи ------------------------------------------------


def test_passport_carries_version_time_event_and_schema_version():
    passport = _passport(_sheet(_build()))
    assert passport["Версия"] == 3
    assert passport["Время сдачи"] == "04.06.2026 17:05"
    assert passport["Событие"] == "С изменениями"
    assert passport["Версия схемы снапшота"] == 1
    assert passport["Подразделение"] == "Управление А"
    assert passport["Сдал"] == "seed-op"
    assert passport["Дата"] == "04.06.2026"
    assert passport["Актуальная"] == "Да"
    assert passport["Опоздание"] == "Нет"


def test_submitted_at_cell_is_a_string_not_a_datetime():
    # Ловушка №1: openpyxl не пишет tz-aware datetime в ячейку (ValueError).
    # Паспорт обязан нести ГОТОВУЮ строку.
    value = _passport(_sheet(_build()))["Время сдачи"]
    assert isinstance(value, str)


def test_passport_carries_exactly_the_nine_labels_in_order():
    """AC-2 перечисляет девять строк шапки — пин на состав И порядок.

    Поштучные ассерты выше переживают выпавшую строку, если её никто не
    спрашивает; этот пин не переживает ни пропажу, ни перестановку.
    """
    sheet = _sheet(_build())
    stop = _header_row_index(sheet)
    labels = [
        row[0]
        for row in sheet.iter_rows(
            min_row=1, max_row=stop - 1, max_col=1, values_only=True
        )
        if row[0] is not None
    ]
    assert labels == [
        "Подразделение",
        "Дата",
        "Версия",
        "Событие",
        "Актуальная",
        "Сдал",
        "Время сдачи",
        "Опоздание",
        "Версия схемы снапшота",
    ]


# --- AC-2: приведение времени сдачи к локальной зоне ---------------------


def test_submitted_at_label_converts_utc_into_the_local_timezone():
    """Гвард минусовой... вернее, ПЛЮСОВОЙ зоны: БД хранит UTC.

    Django с ``USE_TZ`` отдаёт ``submitted_at`` в UTC, а «щит» обязан печатать
    местное время. Инстант выбран так, что приведение ПЕРЕВОДИТ ДАТУ через
    полночь (20:30 UTC → 01:30 следующего дня в ``Asia/Qyzylorda``, +05):
    выпавший ``astimezone`` меняет и час, и число, а не только час. Формат-
    регексп в API-тесте такую потерю пережил бы — этот ассерт нет.
    """
    utc_instant = datetime(2026, 6, 4, 20, 30, tzinfo=timezone.utc)
    label = personal_export_service._submitted_at_label(utc_instant)
    assert label == "05.06.2026 01:30"


def test_submitted_at_label_uses_the_configured_local_zone_not_a_hardcoded_offset():
    """Знаменатель гварда выше: зона берётся из настройки, а не зашита числом."""
    local_zone = ZoneInfo(settings.VAPS_LOCAL_TIMEZONE)
    utc_instant = datetime(2026, 6, 4, 20, 30, tzinfo=timezone.utc)
    expected = utc_instant.astimezone(local_zone).strftime("%d.%m.%Y %H:%M")
    assert personal_export_service._submitted_at_label(utc_instant) == expected
    # Вакуум-гвард самой фикстуры: если бы зона была UTC, тест выше ничего
    # не доказывал бы — он совпал бы с непреобразованным значением.
    assert expected != utc_instant.strftime("%d.%m.%Y %H:%M")


def test_division_title_is_printed_verbatim_so_uuid_fallback_survives():
    # Фолбэк «нет имени → сам UUID» живёт в сервисе; билдер обязан напечатать
    # переданное значение дословно, иначе фолбэк не доедет до файла (AC-2).
    title = "1e2d3c4b-0000-0000-0000-00000000abcd"
    assert _passport(_sheet(_build(division_title=title)))["Подразделение"] == title


def test_unknown_event_label_is_printed_as_is():
    assert (
        _passport(_sheet(_build(event_label="WEIRD_EVENT")))["Событие"]
        == "WEIRD_EVENT"
    )


@pytest.mark.parametrize(
    ("event", "expected"),
    [
        ("CONFIRMED_NO_CHANGES", "Без изменений"),
        ("CHANGED", "С изменениями"),
        ("AMENDED", "Исправление"),
        ("NOT_A_REAL_EVENT", "NOT_A_REAL_EVENT"),
    ],
)
def test_event_label_mapping_falls_back_to_the_code(event, expected):
    assert personal_export_service.event_label(event) == expected


def test_late_and_stale_submission_render_negatives():
    payload = _build(late=True, is_current=False)
    passport = _passport(_sheet(payload))
    assert passport["Опоздание"] == "Да"
    assert passport["Актуальная"] == "Нет"


# --- AC-3: таблица состава ----------------------------------------------


def test_two_statuses_yield_two_rows_with_continuous_numbering():
    snapshot = _snapshot(
        rows=[
            _row(EMP_A, "SICK_LEAVE", status_id=1),
            _row(EMP_A, "DUTY", status_id=2),
        ]
    )
    table = _table(_sheet(_build(snapshot=snapshot)))
    assert [row[0] for row in table] == [1, 2, 3]
    assert [row[2] for row in table] == ["Абдиров А.А.", "Абдиров А.А.", "Бекова Б.Б."]
    assert [row[4] for row in table] == ["SICK_LEAVE", "DUTY", None]


def test_employee_without_rows_yields_exactly_one_empty_row():
    table = _table(_sheet(_build(snapshot=_snapshot(rows=[_row(EMP_A, "DUTY")]))))
    empty = [row for row in table if row[2] == "Бекова Б.Б."]
    assert len(empty) == 1
    assert empty[0][3] is None
    assert empty[0][4] is None
    assert empty[0][5] is None
    assert empty[0][6] is None
    assert empty[0][7] is None


def test_status_name_resolves_through_the_catalog_and_code_is_always_printed():
    table = _table(_sheet(_build(snapshot=_snapshot(rows=[_row(EMP_A, "SICK_LEAVE")]))))
    assert table[0][3] == "Больничный"
    assert table[0][4] == "SICK_LEAVE"


def test_unknown_status_code_falls_back_to_the_code_itself():
    snapshot = _snapshot(rows=[_row(EMP_A, "RENAMED_AWAY")])
    table = _table(_sheet(_build(snapshot=snapshot)))
    assert table[0][3] == "RENAMED_AWAY"
    assert table[0][4] == "RENAMED_AWAY"


def test_snapshot_row_carrying_the_derived_looking_code_is_printed_from_catalog():
    # AC-4 тонкость: код-справочник, ПРИШЕДШИЙ строкой снапшота, печатается
    # через общий маппинг — это факт, а не подстановка умолчания.
    snapshot = _snapshot(rows=[_row(EMP_A, "IN_SERVICE")])
    table = _table(_sheet(_build(snapshot=snapshot)))
    assert table[0][3] == "В строю"
    assert table[0][4] == "IN_SERVICE"


def test_dates_are_formatted_and_half_open_end_is_printed_as_is():
    snapshot = _snapshot(
        rows=[_row(EMP_A, "DUTY", start="2026-06-01", end="2026-06-10")]
    )
    table = _table(_sheet(_build(snapshot=snapshot)))
    assert table[0][5] == "01.06.2026"
    # Полуоткрытый интервал: «минус день» НЕ вычитаем.
    assert table[0][6] == "10.06.2026"


def test_unparsable_date_is_printed_as_is():
    snapshot = _snapshot(rows=[_row(EMP_A, "DUTY", end="не-дата")])
    assert _table(_sheet(_build(snapshot=snapshot)))[0][6] == "не-дата"


@pytest.mark.parametrize(
    ("source", "expected"),
    [
        ("USER", "Оператор"),
        ("KU_SYNC", "Синхронизация КУ"),
        ("OM_AUTO", "Автоматически (дежурства)"),
        ("SOMETHING_NEW", "SOMETHING_NEW"),
    ],
)
def test_source_labels_are_russian_with_passthrough_fallback(source, expected):
    row = _row(EMP_A, "DUTY")
    row["source"] = source
    table = _table(_sheet(_build(snapshot=_snapshot(rows=[row]))))
    assert table[0][7] == expected


def test_rows_keep_snapshot_order_without_local_sorting():
    """Порядок — снапшота, а не билдера.

    Фикстура намеренно ПЕРЕВЁРНУТА (C, B, A): и по ``employee_id``, и по ФИО
    она убывает, поэтому любая своя сортировка — по имени, по званию или по
    id — переставит строки и покрасит тест. На возрастающей фикстуре такой
    тест был бы вакуумным: сортировка дала бы тот же порядок.
    """
    roster = [
        _member(EMP_C, "Вагапов В.В.", "сержант"),
        _member(EMP_B, "Бекова Б.Б.", "лейтенант"),
        _member(EMP_A, "Абдиров А.А.", "капитан"),
    ]
    rows = [
        _row(EMP_C, "DUTY", status_id=5),
        _row(EMP_B, "SICK_LEAVE", status_id=2),
        _row(EMP_A, "DUTY", status_id=7),
    ]
    table = _table(_sheet(_build(snapshot=_snapshot(roster=roster, rows=rows))))
    assert [row[2] for row in table] == [
        "Вагапов В.В.",
        "Бекова Б.Б.",
        "Абдиров А.А.",
    ]


def test_multiple_rows_of_one_employee_keep_their_snapshot_order():
    """Внутри сотрудника порядок ``rows`` тоже не переставляется.

    ``status_id`` убывает (9 → 2), коды идут против алфавита — своя сортировка
    по любому из этих ключей перевернула бы пару.
    """
    rows = [
        _row(EMP_A, "SICK_LEAVE", status_id=9),
        _row(EMP_A, "DUTY", status_id=2),
    ]
    table = _table(_sheet(_build(snapshot=_snapshot(rows=rows))))
    assert [row[4] for row in table[:2]] == ["SICK_LEAVE", "DUTY"]


# --- AC-4: никакой деривации, полнота состава ---------------------------


def test_orphan_row_outside_roster_is_printed_without_crashing():
    snapshot = _snapshot(rows=[_row(ORPHAN, "DUTY")])
    table = _table(_sheet(_build(snapshot=snapshot)))
    orphan = [row for row in table if row[4] == "DUTY"]
    assert len(orphan) == 1
    assert orphan[0][2] in (None, "")
    assert orphan[0][1] in (None, "")


def test_completeness_equality_on_a_fixture_without_orphans():
    snapshot = _snapshot(
        rows=[_row(EMP_A, "DUTY", status_id=1), _row(EMP_A, "SICK_LEAVE", status_id=2)]
    )
    payload = _build(snapshot=snapshot)
    sheet = _sheet(payload)
    printed = {row[2] for row in _table(sheet)}
    assert printed == {"Абдиров А.А.", "Бекова Б.Б."}


def test_completeness_superset_and_per_employee_row_count():
    """AC-4 гвард полноты: надмножество + max(len(rows), 1) на сотрудника."""
    roster = [
        _member(EMP_A, "Абдиров А.А.", "капитан"),
        _member(EMP_B, "Бекова Б.Б.", "лейтенант"),
        _member(EMP_C, "Вагапов В.В.", "сержант"),
    ]
    rows = [
        _row(EMP_A, "DUTY", status_id=1),
        _row(EMP_A, "SICK_LEAVE", status_id=2),
        _row(EMP_C, "DUTY", status_id=3),
        _row(ORPHAN, "DUTY", status_id=4),
    ]
    snapshot = _snapshot(roster=roster, rows=rows)
    table = _table(_sheet(_build(snapshot=snapshot)))

    names = [row[2] for row in table]
    roster_names = {member["full_name"] for member in roster}
    # Надмножество, не равенство: сирота легальна и тоже печатается.
    assert roster_names <= set(names)
    assert len(table) == 5
    counts = {name: names.count(name) for name in roster_names}
    assert counts == {"Абдиров А.А.": 2, "Бекова Б.Б.": 1, "Вагапов В.В.": 1}


def test_footer_carries_roster_total_and_empty_status_legend():
    snapshot = _snapshot(rows=[_row(EMP_A, "DUTY")])
    footer = _footer_text(_sheet(_build(snapshot=snapshot)))
    assert "Всего в списочном составе: 2" in footer
    assert "Пустой статус — в снапшоте статуса нет" in footer


def test_roster_total_counts_roster_not_status_rows():
    roster = [_member(EMP_A, "Абдиров А.А.", "капитан")]
    rows = [
        _row(EMP_A, "DUTY", status_id=1),
        _row(EMP_A, "SICK_LEAVE", status_id=2),
        _row(ORPHAN, "DUTY", status_id=3),
    ]
    footer = _footer_text(_sheet(_build(snapshot=_snapshot(roster=roster, rows=rows))))
    assert "Всего в списочном составе: 1" in footer


def test_empty_roster_still_renders_a_book():
    payload = _build(snapshot=_snapshot(roster=[], rows=[]))
    sheet = _sheet(payload)
    assert _table(sheet) == []
    assert "Всего в списочном составе: 0" in _footer_text(sheet)


def test_export_modules_carry_no_derived_default():
    """Анти-реинвент-гвард (AC-4): подстановки умолчания в ИСХОДНИКЕ нет."""
    for module in (personal_export, personal_export_service):
        source = Path(module.__file__).read_text(encoding="utf-8")
        assert "IN_SERVICE" not in source, module.__name__
        assert "В строю" not in source, module.__name__


def test_export_modules_do_not_import_derivation_machinery():
    for module in (personal_export, personal_export_service):
        source = Path(module.__file__).read_text(encoding="utf-8")
        for banned in (
            "derive_report",
            "resolve_status",
            "build_expense_document",
            "StrengthReportService",
            "EmployeeStatus",
        ):
            assert banned not in source, f"{module.__name__}: {banned}"


# --- вакуум-гвард --------------------------------------------------------


def test_payload_is_real_bytes_not_an_empty_buffer():
    payload = _build()
    assert payload
    assert isinstance(payload, bytes)
    assert len(payload) > 1000


def test_sheet_title_is_the_iso_business_date():
    assert _sheet(_build()).title == "2026-06-04"
