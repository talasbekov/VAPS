"""Story 10.8 — HTTP-контракт личного экспорта («щит») + обязательный аудит.

Доказывает СКВОЗНУЮ цепочку, а не мок: реальные сотрудники и статусы →
реальный ``submit_day`` → ``GET /{pk}/export/`` → байты открываются openpyxl
и несут ФИО из снапшота и имя типа статуса из справочника.

Порядок гвардов зеркалит ``retrieve``: ``by_id`` → 404 → скоуп. Отдельно
пришпилен GET-специфический случай ``POST /export/`` ⇒ 405 (а не 403):
``post`` ЕСТЬ в ``http_method_names``, поэтому запрос доходит до миксина и
попадает в ветку ``self.action is None → MethodNotAllowed``.

Аудит — обязательный: успех даёт РОВНО одну строку ``SUBMISSION_EXPORTED`` с
``entity_id = division_id`` (не pk сдачи — Ловушка №6: ``UUIDField`` молча
превратил бы int в правдоподобный мусор), отказ по схеме не даёт ни одной.
"""

import io
import itertools
import re
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import pytest
from django.conf import settings
from django.core.management import call_command
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.audit.models import AuditLog
from apps.core import clock
from apps.core.models import (
    Division,
    DivisionType,
    Employee,
    Organization,
)
from apps.operations.rbac.models import UserRole
from apps.operations.statuses.models import EmployeeStatus
from apps.operations.submissions.services import (
    amend_day,
    personal_export_service,
    submit_day,
)

pytestmark = pytest.mark.django_db

TODAY = date(2026, 6, 4)
XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_iin_seq = itertools.count(900)


@pytest.fixture(autouse=True)
def frozen_clock():
    with clock.override(TODAY):
        yield


@pytest.fixture
def tree():
    """seed_operations + seed_statuses, корень и чужое подразделение."""
    call_command("seed_operations")
    call_command("seed_statuses")
    org = Organization.objects.create(name="HQ", code="HQ-PE")
    dt = DivisionType.objects.get_or_create(
        code="management", defaults={"name": "Управление"}
    )[0]
    root = Division.objects.create(
        organization=org, type_code=dt, name="Управление А", code="R-PE"
    )
    other = Division.objects.create(
        organization=org, type_code=dt, name="Чужое", code="O-PE"
    )
    return root, other


@pytest.fixture
def scoped_op(tree):
    root, _ = tree
    UserRole.objects.create(
        user_id="op-scoped", role_code_id="DIVISION_OPERATOR", scope_division_id=root.id
    )
    return "op-scoped"


@pytest.fixture
def other_op(tree):
    """Оператор со скоупом на ЧУЖОЕ подразделение — проба скоуп-гварда."""
    _, other = tree
    UserRole.objects.create(
        user_id="op-other", role_code_id="DIVISION_OPERATOR", scope_division_id=other.id
    )
    return "op-other"


def _make_employee(division, full_name, rank_code=""):
    n = next(_iin_seq)
    return Employee.objects.create(
        iin=f"{n:012d}",
        full_name=full_name,
        rank_code=rank_code,
        position_code="",
        division=division,
        employment_status="WORKING",
    )


def _client(actor="op-scoped"):
    c = APIClient()
    if actor is not None:
        c.credentials(HTTP_X_USER_ID=actor)
    return c


def _url(pk):
    return reverse("ops-daily-submission-export", kwargs={"pk": str(pk)})


def _submitted(division, business_date=TODAY):
    return submit_day(
        division_id=division.id, business_date=business_date, actor="seed-op"
    )


def _sheet(response):
    return load_workbook(io.BytesIO(response.content)).active


def _column_a_to_b(sheet):
    """{лейбл: значение} по колонкам A/B — паспорт сдачи."""
    return {
        row[0]: row[1]
        for row in sheet.iter_rows(max_col=2, values_only=True)
        if row[0] is not None
    }


def _cells(sheet):
    return {
        value
        for row in sheet.iter_rows(values_only=True)
        for value in row
        if value is not None
    }


# --- AC-1/AC-7: успешная выдача -----------------------------------------


def test_export_returns_xlsx_bytes(scoped_op, tree):
    root, _ = tree
    submission = _submitted(root)
    resp = _client(scoped_op).get(_url(submission.pk))
    assert resp.status_code == 200
    assert resp["Content-Type"] == XLSX
    assert resp.content


def test_content_disposition_carries_utf8_filename_star(scoped_op, tree):
    root, _ = tree
    submission = _submitted(root)
    disposition = _client(scoped_op).get(_url(submission.pk))["Content-Disposition"]
    # Фронт (10.8a) читает ИСКЛЮЧИТЕЛЬНО filename*=; ASCII-имя дало бы null.
    assert "filename*=utf-8''" in disposition
    # Кириллица доезжает percent-encoded — «сдача_» в UTF-8.
    assert "%D1%81%D0%B4%D0%B0%D1%87%D0%B0" in disposition
    assert ".xlsx" in disposition
    assert f"v{submission.version}" in disposition


def test_body_is_a_real_workbook_carrying_version_and_submitted_at(scoped_op, tree):
    root, _ = tree
    submission = _submitted(root)
    resp = _client(scoped_op).get(_url(submission.pk))
    passport = _column_a_to_b(_sheet(resp))
    assert passport["Версия"] == submission.version
    assert passport["Версия схемы снапшота"] == 1
    assert passport["Подразделение"] == "Управление А"
    assert passport["Сдал"] == "seed-op"
    # Время сдачи — ГОТОВАЯ строка (openpyxl не пишет tz-aware в ячейку).
    assert isinstance(passport["Время сдачи"], str)
    assert re.fullmatch(r"\d{2}\.\d{2}\.\d{4} \d{2}:\d{2}", passport["Время сдачи"])


def test_submitted_at_is_printed_in_local_time_not_the_utc_stored_in_the_db(
    scoped_op, tree
):
    """Сквозной гвард зоны: БД хранит UTC, файл обязан нести местное время.

    Инстант подобран так, что приведение к ``Asia/Qyzylorda`` (+05) переводит
    дату через полночь: 04.06 20:30 UTC → 05.06 01:30. Формат-регексп
    (``test_body_is_a_real_workbook_...``) выпавший ``astimezone`` пережил бы —
    строка осталась бы правильной ФОРМЫ, но неправильного ЧАСА и ЧИСЛА.
    """
    root, _ = tree
    submission = _submitted(root)
    submission.submitted_at = datetime(2026, 6, 4, 20, 30, tzinfo=timezone.utc)
    submission.save(update_fields=["submitted_at"])

    passport = _column_a_to_b(_sheet(_client(scoped_op).get(_url(submission.pk))))
    assert passport["Время сдачи"] == "05.06.2026 01:30"


def test_status_catalog_name_travels_from_the_selector_into_the_book(scoped_op, tree):
    """Русское имя типа статуса доезжает из справочника в книгу.

    Ассерт именно на ИМЯ («На дежурстве»), а не на код: код печатается в
    отдельной колонке всегда и попадает в файл даже с пустым справочником —
    поэтому `"DUTY" in cells` не доказывает, что ``StatusTypeSelector``
    вообще подключён. С ``status_names={}`` этот тест краснеет, тот — нет.
    """
    root, _ = tree
    employee = _make_employee(root, "Абдиров Асхат Асанович")
    EmployeeStatus.objects.create(
        employee_id=employee.id,
        status_type_code="DUTY",
        date_start=TODAY,
        date_end=TODAY + timedelta(days=2),
        source="USER",
    )
    submission = _submitted(root)
    cells = _cells(_sheet(_client(scoped_op).get(_url(submission.pk))))

    assert "На дежурстве" in cells
    assert "DUTY" in cells


def test_roster_and_status_name_travel_end_to_end(scoped_op, tree):
    """Сквозная цепочка: сотрудник + статус → снапшот → байты книги."""
    root, _ = tree
    employee = _make_employee(root, "Абдиров Асхат Асанович")
    EmployeeStatus.objects.create(
        employee_id=employee.id,
        status_type_code="DUTY",
        date_start=TODAY,
        date_end=TODAY + timedelta(days=2),
        source="USER",
    )
    silent = _make_employee(root, "Бекова Бота Бекқызы")

    submission = _submitted(root)
    cells = _cells(_sheet(_client(scoped_op).get(_url(submission.pk))))

    assert "Абдиров Асхат Асанович" in cells
    # Сотрудник БЕЗ статусов обязан присутствовать — знаменатель сходится.
    assert silent.full_name in cells
    # Имя типа статуса разрешено через справочник, код печатается всегда.
    assert "DUTY" in cells
    assert "Оператор" in cells
    assert "Всего в списочном составе: 2" in cells


def test_division_title_falls_back_to_uuid_when_absent_from_the_map(
    scoped_op, tree
):
    """AC-2 фолбэк: подразделения нет в выборке ⇒ печатается сам UUID."""
    root, _ = tree
    submission = _submitted(root)
    division_id = root.id
    Division.objects.filter(pk=division_id).delete()
    passport = _column_a_to_b(_sheet(_client(scoped_op).get(_url(submission.pk))))
    assert passport["Подразделение"] == str(division_id)


def test_export_returns_the_requested_version_not_the_chain_head(scoped_op, tree):
    root, _ = tree
    stale = _submitted(root)
    amend_day(
        division_id=root.id,
        business_date=TODAY,
        actor="seed-op",
        reason="уточнение состава",
        sanction="замечание",
    )
    stale.refresh_from_db()
    assert stale.is_current is False

    passport = _column_a_to_b(_sheet(_client(scoped_op).get(_url(stale.pk))))
    assert passport["Версия"] == stale.version
    assert passport["Актуальная"] == "Нет"


# --- AC-1: гейт и порядок гвардов ---------------------------------------


def test_anonymous_403(tree):
    root, _ = tree
    submission = _submitted(root)
    resp = _client(actor=None).get(_url(submission.pk))
    assert resp.status_code == 403
    assert resp.data["error_code"] == "PERMISSION_DENIED"


def test_foreign_division_403(other_op, tree):
    root, _ = tree
    submission = _submitted(root)
    resp = _client(other_op).get(_url(submission.pk))
    assert resp.status_code == 403
    assert resp.data["error_code"] == "PERMISSION_DENIED"
    assert resp.data["details"]["division_id"] == str(root.id)


def test_missing_pk_404(scoped_op, tree):
    resp = _client(scoped_op).get(_url(999999))
    assert resp.status_code == 404
    assert resp.data["error_code"] == "ENTITY_NOT_FOUND"


def test_pk_resolves_before_scope_so_phantom_pk_is_404_not_403(other_op, tree):
    """Доказывает порядок by_id → 404 → скоуп (а не скоуп → 403)."""
    resp = _client(other_op).get(_url(999999))
    assert resp.status_code == 404
    assert resp.data["error_code"] == "ENTITY_NOT_FOUND"


@pytest.mark.parametrize("method", ["put", "patch", "delete"])
def test_write_verbs_405(scoped_op, tree, method):
    root, _ = tree
    submission = _submitted(root)
    assert getattr(_client(scoped_op), method)(_url(submission.pk)).status_code == 405


@pytest.mark.parametrize("actor", ["op-scoped", None])
def test_post_on_export_is_405_not_403(scoped_op, tree, actor):
    """`post` ЕСТЬ в http_method_names ⇒ action=None ⇒ 405, не мнимый 403.

    403 вместо 405 означал бы, что гейт сработал раньше маршрутизации.
    """
    root, _ = tree
    submission = _submitted(root)
    assert _client(actor).post(_url(submission.pk)).status_code == 405


# --- AC-5: schema-guard --------------------------------------------------


@pytest.mark.parametrize("broken", [999, "1", True, None])
def test_unsupported_snapshot_schema_422_without_audit(scoped_op, tree, broken):
    root, _ = tree
    submission = _submitted(root)
    if broken is None:
        submission.snapshot.pop("schema_version")
    else:
        submission.snapshot["schema_version"] = broken
    submission.save(update_fields=["snapshot"])

    resp = _client(scoped_op).get(_url(submission.pk))
    assert resp.status_code == 422
    assert resp.data["error_code"] == "SNAPSHOT_SCHEMA_UNSUPPORTED"
    assert resp.data["details"]["supported"] == 1
    # Гвард стоит ДО генерации: журнал выгрузки пуст (посев submit_day свои
    # строки уже написал — считаем именно экспортные).
    assert AuditLog.objects.filter(action="SUBMISSION_EXPORTED").count() == 0
    # AC-5 требует в details ОБА ключа: `supported` без `schema_version` не
    # говорит, ЧТО именно приехало неподдерживаемого.
    assert resp.data["details"]["schema_version"] == repr(
        None if broken is None else broken
    )


def test_schema_guard_refuses_before_generating_any_bytes(scoped_op, tree):
    """Отказ по схеме отдаёт конверт ошибки, а не книгу.

    Пин на то, что 422 — это JSON-ошибка, а не «сгенерировали и передумали»:
    xlsx-байтов в ответе быть не должно.
    """
    root, _ = tree
    submission = _submitted(root)
    submission.snapshot["schema_version"] = 999
    submission.save(update_fields=["snapshot"])

    resp = _client(scoped_op).get(_url(submission.pk))
    assert resp.status_code == 422
    assert XLSX not in resp["Content-Type"]
    # PK-сигнатура zip/xlsx — её в теле отказа быть не может.
    assert not resp.content.startswith(b"PK")


# --- AC-6: обязательный аудит -------------------------------------------


def test_successful_export_writes_exactly_one_audit_row(scoped_op, tree):
    root, _ = tree
    employee = _make_employee(root, "Абдиров Асхат Асанович")
    EmployeeStatus.objects.create(
        employee_id=employee.id,
        status_type_code="DUTY",
        date_start=TODAY,
        date_end=TODAY + timedelta(days=2),
        source="USER",
    )
    submission = _submitted(root)

    resp = _client(scoped_op).get(_url(submission.pk), HTTP_X_REQUEST_ID="req-10-8")
    assert resp.status_code == 200

    logs = list(AuditLog.objects.filter(action="SUBMISSION_EXPORTED"))
    assert len(logs) == 1
    log = logs[0]
    assert log.entity_type == "daily_submission"
    # Ловушка №6: entity_id — UUID подразделения, НЕ целый pk сдачи (запись
    # int'а не падает, она молча оставляет правдоподобный мусор).
    assert log.entity_id == root.id
    assert log.actor_user_id == "op-scoped"
    assert log.request_id == "req-10-8"
    assert log.new_value["submission_id"] == submission.pk
    assert log.new_value["division_id"] == str(root.id)
    assert log.new_value["business_date"] == TODAY.isoformat()
    assert log.new_value["version"] == submission.version
    assert log.new_value["is_current"] is True
    assert log.new_value["event"] == submission.event
    assert log.new_value["roster_size"] == 1
    assert log.new_value["row_count"] == 1
    assert log.new_value["filename"].endswith(".xlsx")


def test_audit_entity_id_is_not_the_submission_pk(scoped_op, tree):
    """Явный анти-гвард Ловушки №6 — int-pk в UUID-поле не отличим на глаз."""
    import uuid

    root, _ = tree
    submission = _submitted(root)
    assert _client(scoped_op).get(_url(submission.pk)).status_code == 200
    log = AuditLog.objects.get(action="SUBMISSION_EXPORTED")
    assert log.entity_id != uuid.UUID(int=submission.pk)


def test_submission_exported_is_declared_in_the_registry():
    path = Path(settings.BASE_DIR).parent.parent / "docs/registries/audit-events.yaml"
    text = path.read_text(encoding="utf-8")
    block = re.search(r"^  SUBMISSION_EXPORTED:\n((?:    .*\n)+)", text, re.M)
    assert block, "SUBMISSION_EXPORTED missing from audit-events.yaml"
    assert "entity_type: daily_submission" in block.group(1)


def test_audit_counters_are_not_interchangeable(scoped_op, tree):
    """``roster_size`` и ``row_count`` — РАЗНЫЕ числа на этой фикстуре.

    Три сотрудника, статусы у двух ⇒ 3 ≠ 2. На симметричной фикстуре (1 и 1)
    перепутанные местами поля дали бы зелёный тест — знаменатель обязан
    отличаться от числителя.
    """
    root, _ = tree
    for name in ("Абдиров А.А.", "Бекова Б.Б."):
        employee = _make_employee(root, name)
        EmployeeStatus.objects.create(
            employee_id=employee.id,
            status_type_code="DUTY",
            date_start=TODAY,
            date_end=TODAY + timedelta(days=2),
            source="USER",
        )
    _make_employee(root, "Вагапов В.В.")
    submission = _submitted(root)

    assert _client(scoped_op).get(_url(submission.pk)).status_code == 200
    log = AuditLog.objects.get(action="SUBMISSION_EXPORTED")
    assert log.new_value["roster_size"] == 3
    assert log.new_value["row_count"] == 2


def test_audit_records_the_exact_canonical_filename(scoped_op, tree):
    """AC-7: имя файла — ``сдача_{ISO}_v{version}.xlsx`` целиком.

    ``endswith(".xlsx")`` пережил бы и потерю даты, и потерю версии.
    """
    root, _ = tree
    submission = _submitted(root)
    assert _client(scoped_op).get(_url(submission.pk)).status_code == 200
    log = AuditLog.objects.get(action="SUBMISSION_EXPORTED")
    assert log.new_value["filename"] == f"сдача_{TODAY.isoformat()}_v1.xlsx"


def test_export_operation_is_declared_in_the_committed_schema():
    """AC-8: ``@extend_schema`` на месте — и в схеме не докстринг миксина.

    ``test_schema_drift`` доказывает лишь, что ``schema.yaml`` совпадает с
    генерацией: снятый ``@extend_schema`` пройдёт его насквозь, если файл
    перегенерировали. Наблюдаемая на ``amend`` болезнь — drf-spectacular
    затягивает в описание докстринг ``RequirePermissionMixin`` — ловится
    только пином на само содержимое операции.
    """
    text = (Path(settings.BASE_DIR) / "schema.yaml").read_text(encoding="utf-8")
    block = re.search(
        r"^  /api/operations/daily-submissions/\{id\}/export/:\n"
        r"((?:(?:    .*)?\n)+?)(?=^  \S)",
        text,
        re.M,
    )
    assert block, "роут export отсутствует в schema.yaml — забыт `make schema`"
    operation = block.group(1)
    assert XLSX in operation, "в схеме нет xlsx-типа содержимого"
    assert "format: binary" in operation
    # Описание — наше, а не подтянутый докстринг миксина.
    assert "RequirePermissionMixin" not in operation
    assert "Личная копия сданного дня" in operation


def test_audit_write_failure_denies_the_file_instead_of_serving_it_silently(
    scoped_op, tree, monkeypatch
):
    """AC-6 «нет журнала ⇒ нет выдачи»: падение ``record`` НЕ глушится.

    Сервис намеренно не оборачивает запись ни в ``atomic``, ни в
    ``try/except``, поэтому падение доходит до общего обработчика и даёт 500.
    Оба ассерта нужны: код 500 доказывает, что глушителя нет, а отсутствие
    zip-сигнатуры — что клиенту не уехал файл БЕЗ строки в журнале (байты к
    этому моменту уже сгенерированы — соблазн «отдать всё равно» реален).
    """
    root, _ = tree
    submission = _submitted(root)

    def _boom(**kwargs):
        raise RuntimeError("аудит недоступен")

    monkeypatch.setattr(personal_export_service, "record", _boom)
    resp = _client(scoped_op).get(_url(submission.pk))

    assert resp.status_code == 500
    assert not resp.content.startswith(b"PK")
    assert AuditLog.objects.filter(action="SUBMISSION_EXPORTED").count() == 0
