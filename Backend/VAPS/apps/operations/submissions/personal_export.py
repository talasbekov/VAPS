"""Story 10.8 — чистый билдер личной копии сданного дня («щит»).

``build_personal_export_xlsx`` собирает из снапшота сдачи (schema v1) книгу
.xlsx: блок-паспорт сдачи (подразделение, дата, версия, событие, кто и когда
сдал, версия схемы) плюс поимённая ведомость ``roster ⋈ rows``. Файл —
ЗЕРКАЛО снапшота: он ничего не выводит и не досчитывает.

Контракт генераторов (6.4): ЧИСТАЯ функция — ни ORM, ни ``Clock``, ни диска,
ни сети. Всё, чего в снапшоте нет (имя подразделения, версия сдачи, время
сдачи, флаги ``is_current``/``late``, справочные имена типов статусов),
приходит ГОТОВЫМИ параметрами — их резолвит сервис-обвязка. ``openpyxl``
импортируется ЛЕНИВО внутри функции (канон ``generate_expense_xlsx``).

Ведомость строится по паре (сотрудник, статус): сотрудник с двумя статусами
даёт две строки, сотрудник БЕЗ статусов — ровно одну с пустыми статусными
ячейками (знаменатель обязан быть виден). Строка, чей ``employee_id`` вне
``roster``, тоже печатается — с пустыми ФИО/званием, без падения. Порядок —
как в снапшоте: ``roster`` уже отсортирован по ``employee_id``, ``rows`` — по
``(employee_id, status_id)``; своей сортировки билдер НЕ заводит.

Даты в ``rows`` — ISO-строки; интервал полуоткрытый ``[date_start, date_end)``
и печатается КАК ЕСТЬ («минус день» не вычитаем). Непарсящееся значение
выводится дословно, а не роняет генерацию.
"""

import io
from datetime import date

_DATE_FORMAT = "%d.%m.%Y"
_YES = "Да"
_NO = "Нет"

TABLE_COLUMNS = (
    "№",
    "Звание",
    "ФИО",
    "Статус",
    "Код статуса",
    "С",
    "По",
    "Источник",
)

# Три значения источника факта — закрытый словарь, отдельный селектор ради
# него не заводим. Незнакомое значение печатается само (тот же принцип, что
# у имени типа статуса): файл обязан пережить рост словаря, а не упасть.
SOURCE_LABELS = {
    "USER": "Оператор",
    "KU_SYNC": "Синхронизация КУ",
    "OM_AUTO": "Автоматически (дежурства)",
}

EMPTY_STATUS_LEGEND = "Пустой статус — в снапшоте статуса нет"
ROSTER_TOTAL_LABEL = "Всего в списочном составе"

_COLUMN_WIDTHS = (6, 16, 32, 24, 22, 13, 13, 26)


def _format_iso_date(value):
    """ISO-строка → ``ДД.ММ.ГГГГ``; непарсящееся значение — дословно."""
    try:
        return date.fromisoformat(value).strftime(_DATE_FORMAT)
    except (TypeError, ValueError):
        return value


def _pair_rows(roster, rows):
    """[(запись roster | {}, строка rows | None)] в порядке снапшота.

    Сотрудник без строк даёт ровно одну пару с ``None``; строки-сироты
    (``employee_id`` вне ``roster``) уходят в хвост с пустой записью roster.
    """
    grouped = {}
    for row in rows:
        grouped.setdefault(str(row.get("employee_id")), []).append(row)

    paired = []
    for member in roster:
        member_rows = grouped.pop(str(member.get("employee_id")), [])
        if member_rows:
            paired.extend((member, row) for row in member_rows)
        else:
            paired.append((member, None))
    for orphan_rows in grouped.values():
        paired.extend(({}, row) for row in orphan_rows)
    return paired


def _table_values(number, member, row, status_names, source_labels):
    """Значения одной строки ведомости в порядке ``TABLE_COLUMNS``."""
    values = [
        number,
        member.get("rank") or "",
        member.get("full_name") or "",
    ]
    if row is None:
        # Статуса в снапшоте нет — ячейки остаются пустыми (AC-4): умолчание
        # сюда не подставляется, производное состояние считает сервер.
        return values + ["", "", "", "", ""]
    code = row.get("status_type_code")
    source = row.get("source")
    return values + [
        status_names.get(code, code),
        code,
        _format_iso_date(row.get("date_start")),
        _format_iso_date(row.get("date_end")),
        source_labels.get(source, source),
    ]


def build_personal_export_xlsx(
    *,
    snapshot,
    division_title,
    business_date,
    version,
    is_current,
    event_label,
    submitted_by,
    submitted_at_label,
    late,
    status_names,
    source_labels=None,
) -> bytes:
    """Снапшот сдачи + атрибуты версии → байты .xlsx личной копии.

    ``submitted_at_label`` — ГОТОВАЯ строка, а не ``datetime``: openpyxl не
    пишет tz-aware значение в ячейку (``ValueError``). Приведение к локальной
    зоне и формат — забота сервиса.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    labels = SOURCE_LABELS if source_labels is None else source_labels
    bold = Font(bold=True)

    workbook = Workbook()
    sheet = workbook.active
    # ISO-дата свободна от запрещённых openpyxl символов ``/\?*[]:``.
    sheet.title = business_date.isoformat()

    roster = snapshot.get("roster") or []
    rows = snapshot.get("rows") or []

    passport = (
        ("Подразделение", division_title),
        ("Дата", business_date.strftime(_DATE_FORMAT)),
        ("Версия", version),
        ("Событие", event_label),
        ("Актуальная", _YES if is_current else _NO),
        ("Сдал", submitted_by),
        ("Время сдачи", submitted_at_label),
        ("Опоздание", _YES if late else _NO),
        ("Версия схемы снапшота", snapshot.get("schema_version")),
    )
    for index, (label, value) in enumerate(passport, start=1):
        sheet.cell(row=index, column=1, value=label).font = bold
        sheet.cell(row=index, column=2, value=value)

    header_row = len(passport) + 2
    for column, label in enumerate(TABLE_COLUMNS, start=1):
        sheet.cell(row=header_row, column=column, value=label).font = bold

    paired = _pair_rows(roster, rows)
    for number, (member, row) in enumerate(paired, start=1):
        values = _table_values(number, member, row, status_names, labels)
        for column, value in enumerate(values, start=1):
            sheet.cell(row=header_row + number, column=column, value=value)

    # Пустая строка-разделитель: подвал не должен читаться как хвост таблицы.
    footer_row = header_row + len(paired) + 2
    sheet.cell(row=footer_row, column=1, value=f"{ROSTER_TOTAL_LABEL}: {len(roster)}")
    sheet.cell(row=footer_row + 1, column=1, value=EMPTY_STATUS_LEGEND)

    for column, width in enumerate(_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[chr(ord("A") + column - 1)].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
